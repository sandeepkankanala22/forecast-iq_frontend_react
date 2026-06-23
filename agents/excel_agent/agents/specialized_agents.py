"""
Specialized agent implementations for Excel creation
"""

import re
import sys
from typing import Dict, List, Optional, Tuple
import subprocess
import ast
import time
import json
from pathlib import Path
from .base_agent import BaseAgent
from ..utils.bedrock_client import BedrockClient
from ..utils.prompt_manager import PromptManager
from ..utils.excel_utils import ExcelManager
from ..utils.code_parser import clean_code_response
from ..utils.logging_utils import get_session_manager

def convert_to_utf8(text):
    """Convert text to UTF-8 encoding."""
    if isinstance(text, bytes):
        return text.decode('utf-8', errors='ignore')
    return str(text)

def parse_json(text):
    """Parse a text into a JSON object (dict or list)

    Args:
        text (String): text to parse

    Returns:
        dict or list: parsed JSON object
    """
    text = convert_to_utf8(text)
    
    # Strip markdown code blocks if present
    # Handle ```json ... ``` or ``` ... ```
    markdown_pattern = r"```(?:json)?\s*(.*?)\s*```"
    markdown_match = re.search(markdown_pattern, text, re.DOTALL)
    if markdown_match:
        text = markdown_match.group(1).strip()
    
    # Try direct JSON parsing first
    try:
        parsed = json.loads(text)
        if isinstance(parsed, (dict, list)):
            return parsed
    except json.JSONDecodeError as e:
        # Log the specific error for debugging
        print(f"JSON parse error: {e}", file=sys.stderr)
        pass
    
    # Try to extract JSON (dict or list) using regex
    # First try arrays
    array_pattern = r"\[.*\]"
    array_matches = re.findall(array_pattern, text, re.DOTALL)
    for match in array_matches:
        try:
            parsed = json.loads(match)
            if isinstance(parsed, list):
                return parsed
        except json.JSONDecodeError:
            pass
    
    # Then try objects
    dict_pattern = r"\{.*\}"
    dict_matches = re.findall(dict_pattern, text, re.DOTALL)
    for match in dict_matches:
        try:
            parsed = json.loads(match)
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            pass
        # Fallback to ast.literal_eval
        try:
            parsed = ast.literal_eval(match)
            if isinstance(parsed, dict):
                return parsed
        except Exception:
            continue
    
    return None

def parse_dict(text):
    """Parse a text into a dictionary

    Args:
        text (String): text to parse

    Returns:
        Dictionary: parsed dictionary
    """
    text = convert_to_utf8(text)
    
    # Strip markdown code blocks if present
    # Handle ```json ... ``` or ``` ... ```
    markdown_pattern = r"```(?:json)?\s*(.*?)\s*```"
    markdown_match = re.search(markdown_pattern, text, re.DOTALL)
    if markdown_match:
        text = markdown_match.group(1).strip()
    
    # Check for incomplete JSON (common issue)
    if text.count('{') != text.count('}'):
        # Try to fix incomplete JSON by adding missing braces
        if text.count('{') > text.count('}'):
            missing_braces = text.count('{') - text.count('}')
            text += '}' * missing_braces
        elif text.count('}') > text.count('{'):
            # Remove extra closing braces
            extra_braces = text.count('}') - text.count('{')
            for _ in range(extra_braces):
                text = text.rsplit('}', 1)[0]
    
    # Try direct JSON parsing first
    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            return parsed
    except json.JSONDecodeError as e:
        # Log the specific error for debugging
        print(f"JSON parse error: {e}", file=sys.stderr)
        pass
    
    # Try to extract JSON using regex
    pattern = r"\{.*\}"
    matches = re.findall(pattern, text, re.DOTALL)
    for match in matches:
        # Try json.loads first (more reliable than ast.literal_eval)
        try:
            parsed = json.loads(match)
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            pass
        # Fallback to ast.literal_eval
        try:
            parsed = ast.literal_eval(match)
            if isinstance(parsed, dict):
                return parsed
        except Exception:
            continue
    return {}


class DecomposerAgent(BaseAgent):
    """Agent that decomposes user requests into actionable tasks"""
    
    def __init__(self, bedrock_client: BedrockClient, prompt_manager: PromptManager, session_manager=None):
        super().__init__(
            name="Decomposer",
            prompt_name="decomposer",
            bedrock_client=bedrock_client,
            prompt_manager=prompt_manager,
            temperature=0.7,
            session_manager=session_manager,
        )
    
    def decompose_request(self, user_request: str) -> Dict:
        """
        Break down user request into structured components

        Args:
            user_request: Original user request

        Returns:
            Dictionary with decomposed tasks including:
            - query_type: Type of query (new/edit/ask)
            - query_decomposition: Expanded description
            - input_files: List of input files or None
            - continuation: Whether this continues previous work
            - original_request: The original user request
            - agent: Name of this agent
        """
        self.logger.info(f"Decomposing request: {user_request[:100]}...")

        system_prompt, user_template = self.prompt_manager.get_prompt_parts("decomposer_decompose_request")
        user_message = user_template.format(user_request=user_request)
        response = self.generate_response(user_message, system_prompt_override=system_prompt)

        # Parse the JSON response
        parsed_response = parse_json(response)

        if parsed_response is None:
            self.logger.warning("Failed to parse JSON response from decomposer, falling back to text response")
            # Fallback to original behavior if parsing fails
            return {
                'original_request': user_request,
                'query_type': 'unknown',
                'query_decomposition': response,
                'input_files': None,
                'continuation': False,
                'agent': self.name,
                'parse_error': True
            }

        # Validate required fields
        required_fields = ['query_type', 'query_decomposition', 'input_files', 'continuation']
        for field in required_fields:
            if field not in parsed_response:
                self.logger.warning(f"Missing required field '{field}' in decomposer response")
                parsed_response[field] = None

        # Add metadata
        parsed_response['original_request'] = user_request
        parsed_response['agent'] = self.name

        self.logger.info(f"Successfully decomposed request: type={parsed_response.get('query_type')}, continuation={parsed_response.get('continuation')}")

        return parsed_response

class CriticAgent(BaseAgent):
    """Agent that reviews and critiques spreadsheet quality"""
    
    def __init__(
        self,
        bedrock_client: BedrockClient,
        prompt_manager: PromptManager,
        excel_manager: ExcelManager,
        session_manager=None,
    ):
        super().__init__(
            name="Critic",
            prompt_name="critic",
            bedrock_client=bedrock_client,
            prompt_manager=prompt_manager,
            temperature=0.5,  # Lower temperature for more focused critique
            session_manager=session_manager,
        )
        self.excel_manager = excel_manager
    
    def review_spreadsheet(
        self,
        filepath: str,
        sheet_name: Optional[str] = None,
        action_plan: Optional[Dict] = None,
        validation_rules: Optional[Dict] = None,
        confidence_thresholds: Optional[Dict] = None,
    ) -> Dict:
        """
        Review a spreadsheet and provide feedback.

        When sheet_name and action_plan are provided, evaluates only that sheet
        against its action plan and requirements, not the whole workbook or total user request.

        Args:
            filepath: Path to Excel file
            sheet_name: Optional name of the sheet to evaluate (evaluate only this sheet)
            action_plan: Optional action plan for that sheet (evaluate against this plan only)

        Returns:
            Dictionary with review results
        """
        self.logger.info(f"Reviewing spreadsheet: {filepath}" + (f" (sheet: {sheet_name})" if sheet_name else ""))

        # Analyze the workbook and extract ACTUAL sheet content for review
        try:
            wb = self.excel_manager.load_workbook(filepath)
            analysis = self.excel_manager.analyze_workbook(wb)
            errors = self.excel_manager.detect_errors(wb)

            # Extract actual sheet content (cells, formulas, grid) for the sheet under review
            sheet_content_str = "N/A"
            if sheet_name and sheet_name in wb.sheetnames:
                content = self.excel_manager.extract_sheet_content(wb, sheet_name)
                if "error" in content:
                    sheet_content_str = content["error"]
                else:
                    parts = [
                        f"Sheet: {content['sheet_name']} | Rows: {content['max_row']} | Cols: {content['max_column']}",
                        f"Sheet names in workbook: {', '.join(content['sheet_names'])}",
                        "\n--- Cell grid (Row N: col values, [F]=formula) ---",
                        content.get("grid_sample", "(no data)"),
                        "\n--- Formulas (check references use valid sheet names and cells) ---",
                    ]
                    for f in content.get("formulas", [])[:80]:
                        parts.append(f"  {f['cell']}: {f['formula']}")
                    if content.get("values_resolved"):
                        parts.append("\n--- Resolved values (errors like #REF! appear here) ---")
                        for coord, val in list(content["values_resolved"].items())[:50]:
                            parts.append(f"  {coord}: {val}")
                    sheet_content_str = "\n".join(parts)

            # When reviewing a single sheet, restrict sheet details to that sheet
            sheets_to_detail = analysis["sheets"]
            if sheet_name and sheet_name in analysis["sheets"]:
                sheets_to_detail = {sheet_name: analysis["sheets"][sheet_name]}

            sheet_under_review = sheet_name if sheet_name else "N/A"
            sheet_action_plan_str = json.dumps(action_plan, indent=2) if action_plan else "N/A"
            validation_rules_str = json.dumps(validation_rules or {}, indent=2)
            confidence_thresholds_str = json.dumps(confidence_thresholds or {}, indent=2)

            system_prompt, user_template = self.prompt_manager.get_prompt_parts("critic_review_spreadsheet")
            user_message = user_template.format(
                sheet_under_review=sheet_under_review,
                sheet_action_plan=sheet_action_plan_str,
                sheet_count=analysis["sheet_count"],
                total_formulas=analysis["total_formulas"],
                total_data_cells=analysis["total_cells_with_data"],
                sheet_details=self._format_sheet_details(sheets_to_detail),
                errors=self._format_errors(errors),
                sheet_content=sheet_content_str,
                validation_rules=validation_rules_str,
                confidence_thresholds=confidence_thresholds_str,
            )
            response = self.generate_response(user_message, system_prompt_override=system_prompt)

            # Parse JSON response
            parsed_review = self._parse_critic_response(response)
            
            return {
                'filepath': filepath,
                'analysis': analysis,
                'errors': errors,
                'feedback': response,  # Keep raw response for logging
                'parsed_review': parsed_review,
                'overall_score': parsed_review.get('overall_score', 0),
                'improvement_points': parsed_review.get('improvement_points', []),
                'summary': parsed_review.get('summary', ''),
                'agent': self.name
            }

        except Exception as e:
            self.logger.error(f"Failed to review spreadsheet", exc_info=e)
            raise

    def _parse_critic_response(self, response: str) -> Dict:
        """
        Parse critic JSON response
        
        Args:
            response: Raw LLM response string
            
        Returns:
            Parsed review dictionary with standardized structure
        """
        try:
            # Try to parse as JSON using the parse_json utility
            parsed = parse_json(response)
            
            if parsed and isinstance(parsed, dict):
                self.logger.info("Successfully parsed critic response as JSON")
                
                # Validate required fields and provide defaults
                standardized = {
                    'overall_score': parsed.get('overall_score', 0),
                    'summary': parsed.get('summary', 'No summary provided'),
                    'strengths': parsed.get('strengths', []),
                    'improvement_points': parsed.get('improvement_points', []),
                    'missing_elements': parsed.get('missing_elements', []),
                    'formula_issues': parsed.get('formula_issues', []),
                    'formatting_issues': parsed.get('formatting_issues', []),
                    'requirements_coverage': parsed.get('requirements_coverage', 'Unknown'),
                    'overall_recommendation': parsed.get('overall_recommendation', '')
                }
                
                # Log if important fields are missing
                if not standardized['improvement_points']:
                    self.logger.warning("No improvement_points found in critic response")
                
                return standardized
            else:
                self.logger.warning("Failed to parse critic response as JSON, using fallback")
                return self._fallback_parse_critic_response(response)
                
        except Exception as e:
            self.logger.error(f"Error parsing critic response: {e}", exc_info=True)
            return self._fallback_parse_critic_response(response)
    
    def _fallback_parse_critic_response(self, response: str) -> Dict:
        """
        Fallback parser for non-JSON critic responses
        
        Args:
            response: Raw text response
            
        Returns:
            Best-effort parsed review dictionary
        """
        self.logger.info("Using fallback text parser for critic response")
        
        # Try to extract a score from text
        score = 0
        score_patterns = [
            r'score[:\s]+(\d+)',
            r'rating[:\s]+(\d+)',
            r'(\d+)\s*/\s*(?:10|100)',
        ]
        for pattern in score_patterns:
            match = re.search(pattern, response, re.IGNORECASE)
            if match:
                score = int(match.group(1))
                break
        
        # Extract improvement points from text (look for bullet points or numbered lists)
        improvement_points = []
        lines = response.split('\n')
        for line in lines:
            line_stripped = line.strip()
            # Look for lines that start with bullets, numbers, or "improvement" keyword
            if (line_stripped.startswith(('-', '*', '•')) or 
                re.match(r'^\d+\.', line_stripped) or
                'improvement' in line_stripped.lower() or
                'fix' in line_stripped.lower() or
                'issue' in line_stripped.lower()):
                
                if len(line_stripped) > 10:  # Avoid very short lines
                    improvement_points.append({
                        'category': 'General',
                        'severity': 'moderate',
                        'location': 'Unknown',
                        'issue_description': line_stripped.lstrip('-*•0123456789. '),
                        'suggested_fix': 'See description'
                    })
        
        return {
            'overall_score': score,
            'summary': 'Fallback parse - JSON not available',
            'strengths': [],
            'improvement_points': improvement_points,
            'missing_elements': [],
            'formula_issues': [],
            'formatting_issues': [],
            'requirements_coverage': 'Unknown',
            'overall_recommendation': response[:200] if len(response) > 200 else response
        }
    
    def _format_sheet_details(self, sheets: Dict, max_sheets: int = 10) -> str:
        """
        Format sheet details for prompt with size limits
        
        Args:
            sheets: Dictionary of sheet info
            max_sheets: Maximum number of sheets to include details for
        
        Returns:
            Formatted string with sheet details (truncated if needed)
        """
        details = []
        sheet_count = 0
        for name, info in sheets.items():
            if sheet_count >= max_sheets:
                remaining = len(sheets) - max_sheets
                details.append(f"... and {remaining} more sheet(s) (truncated for brevity)")
                break
            
            # Limit formula count display for very large sheets
            formula_count = len(info['formulas'])
            formula_text = f"{formula_count} formulas" if formula_count <= 1000 else f"{formula_count} formulas (large sheet)"
            
            details.append(
                f"- {name}: {info['max_row']} rows, {info['max_column']} cols, {formula_text}"
            )
            sheet_count += 1
        return "\n".join(details)
    
    def _format_errors(self, errors: Dict, max_locations: int = 3) -> str:
        """
        Format error details for prompt with size limits
        
        Args:
            errors: Dictionary of error types and locations
            max_locations: Maximum number of error locations to show per type
        
        Returns:
            Formatted string with error details (truncated if needed)
        """
        error_list = []
        for error_type, locations in errors.items():
            if locations:
                error_list.append(f"- {error_type}: {len(locations)} occurrences")
                if len(locations) <= max_locations:
                    error_list.append(f"  Locations: {', '.join(locations)}")
                else:
                    sample_locations = ', '.join(locations[:max_locations])
                    error_list.append(f"  Sample locations: {sample_locations} (and {len(locations) - max_locations} more)")
        return "\n".join(error_list) if error_list else "No errors detected"

class ActionItemsAgent(BaseAgent):
    """Agent that creates actionable implementation plan for a single sheet from user query and assumptions"""

    def __init__(
        self,
        bedrock_client: BedrockClient,
        prompt_manager: PromptManager,
        session_manager=None,
    ):
        super().__init__(
            name="ActionItems",
            prompt_name="action_items",
            bedrock_client=bedrock_client,
            prompt_manager=prompt_manager,
            temperature=0.4,  # Moderate creativity for implementation planning
            session_manager=session_manager,
        )

    def generate_sheet_action_plan(
        self, 
        user_query: str, 
        assumptions_without_sheet_format: Dict, 
        sheet_definition: Dict,
        sheet_index: int
    ) -> Dict:
        """
        Generate actionable implementation plan for a single sheet

        Args:
            user_query: Original user request/query (not decomposed)
            assumptions_without_sheet_format: Assumptions config without sheet_format_defaults key
            sheet_definition: Single sheet definition from sheet_format_defaults.sheets array
            sheet_index: Index of this sheet in the workbook (0-based)

        Returns:
            Dictionary with detailed action plan for this specific sheet
        """
        self.logger.info(f"Generating action plan for sheet: {sheet_definition.get('name', 'Unknown')}")

        try:
            system_prompt, user_template = self.prompt_manager.get_prompt_parts("action_items_generate")
            user_message = user_template.format(
                user_query=user_query,
                assumptions_config=json.dumps(assumptions_without_sheet_format, indent=2),
                sheet_definition=json.dumps(sheet_definition, indent=2),
                sheet_index=sheet_index
            )
            response = self.generate_response(user_message, system_prompt_override=system_prompt)

            # Parse JSON response
            try:
                action_plan = parse_dict(response)
                self.logger.info(f"Successfully generated action plan for sheet '{sheet_definition.get('name')}'")
                return action_plan

            except json.JSONDecodeError as e:
                self.logger.error(f"Failed to parse sheet action plan JSON: {e}")
                self.logger.debug(f"Response was: {response[:500]}")
                return self._get_default_sheet_action_plan(sheet_definition, sheet_index)

        except Exception as e:
            self.logger.error(f"Failed to generate sheet action plan", exc_info=e)
            return self._get_default_sheet_action_plan(sheet_definition, sheet_index)

    def generate_modification_action_plan(
        self,
        user_query: str,
        sheet_name: str,
        existing_code: Optional[str] = None,
        sheet_data: Optional[Dict] = None,
        workbook_path: Optional[str] = None
    ) -> Dict:
        """
        Generate action plan for modifying an existing sheet.
        
        Supports three modes:
        1. Modify with script: When existing_code is provided
        2. Modify from sheet data: When sheet_data is provided (no script)
        3. Hybrid: When both are provided
        
        Args:
            user_query: User's modification request
            sheet_name: Name of the sheet to modify
            existing_code: Optional current code for the sheet (if available)
            sheet_data: Optional current sheet data/structure (if no code or for reference)
            workbook_path: Optional path to workbook (to extract sheet data if needed)
            
        Returns:
            Dictionary with modification action plan (structured for both creation and modification)
        """
        self.logger.info(f"Generating modification action plan for sheet: {sheet_name}")
        
        # Determine mode and load appropriate prompt
        if existing_code:
            mode = "modify_with_script"
            prompt_name = "action_items_modify_with_script"
            self.logger.info("Using modify-with-script mode")
        elif sheet_data:
            mode = "modify_from_data"
            prompt_name = "action_items_modify_from_sheet"
            self.logger.info("Using modify-from-data mode (no script available)")
        elif workbook_path:
            # Try to extract sheet data from workbook
            mode = "modify_from_data"
            prompt_name = "action_items_modify_from_sheet"
            self.logger.info("Extracting sheet data from workbook")
            sheet_data = self._extract_sheet_data(workbook_path, sheet_name)
        else:
            self.logger.warning("No code or data provided for modification, using minimal info")
            return self._get_default_modification_plan(sheet_name, user_query)
        
        try:
            system_prompt, user_template = self.prompt_manager.get_prompt_parts(prompt_name)
            
            # Format sheet_data for the prompt
            sheet_data_str = json.dumps(sheet_data, indent=2) if sheet_data else "Not available"
            existing_code_str = existing_code if existing_code else "Not available"
            
            user_message = user_template.format(
                user_query=user_query,
                sheet_name=sheet_name,
                existing_code=existing_code_str,
                sheet_data=sheet_data_str
            )
            
            response = self.generate_response(user_message, system_prompt_override=system_prompt)
            
            try:
                action_plan = parse_dict(response)
                # Ensure mode is set in the response
                if 'sheet_info' in action_plan:
                    action_plan['sheet_info']['mode'] = mode
                self.logger.info(f"Successfully generated modification action plan for sheet '{sheet_name}' (mode: {mode})")
                return action_plan
            except json.JSONDecodeError as e:
                self.logger.error(f"Failed to parse modification action plan JSON: {e}")
                self.logger.debug(f"Response was: {response[:500]}")
                return self._get_default_modification_plan(sheet_name, user_query)
                
        except Exception as e:
            self.logger.error(f"Failed to generate modification action plan", exc_info=e)
            return self._get_default_modification_plan(sheet_name, user_query)
    
    def _extract_sheet_data(self, workbook_path: str, sheet_name: str) -> Optional[Dict]:
        """
        Extract current data and structure from a sheet in an existing workbook
        
        Args:
            workbook_path: Path to the workbook
            sheet_name: Name of the sheet to extract
            
        Returns:
            Dictionary with sheet structure, data, formulas, formatting (or None if failed)
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(workbook_path, data_only=False)
            
            if sheet_name not in wb.sheetnames:
                self.logger.warning(f"Sheet '{sheet_name}' not found in workbook")
                return None
            
            ws = wb[sheet_name]
            
            # Extract basic info
            sheet_data = {
                "name": sheet_name,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "dimensions": ws.dimensions,
                "cells_sample": [],
                "formulas": [],
                "merged_cells": [str(r) for r in ws.merged_cells.ranges],
            }
            
            # Sample cells (first 100 non-empty cells for context)
            cell_count = 0
            for row in ws.iter_rows(max_row=min(50, ws.max_row)):
                for cell in row:
                    if cell.value is not None and cell_count < 100:
                        cell_info = {
                            "cell": cell.coordinate,
                            "value": str(cell.value)[:100],  # Truncate long values
                            "type": "formula" if isinstance(cell.value, str) and cell.value.startswith('=') else "data"
                        }
                        sheet_data["cells_sample"].append(cell_info)
                        
                        if cell_info["type"] == "formula":
                            sheet_data["formulas"].append({
                                "cell": cell.coordinate,
                                "formula": cell.value
                            })
                        cell_count += 1
            
            self.logger.info(f"Extracted data from sheet '{sheet_name}': {ws.max_row} rows, {ws.max_column} cols, {len(sheet_data['formulas'])} formulas")
            return sheet_data
            
        except Exception as e:
            self.logger.error(f"Failed to extract sheet data from '{workbook_path}': {e}", exc_info=e)
            return None
    
    def _get_default_modification_plan(self, sheet_name: str, user_query: str) -> Dict:
        """
        Get default modification action plan as fallback.
        Returns structure consistent with full action plans.
        
        Args:
            sheet_name: Name of the sheet
            user_query: User's modification request
            
        Returns:
            Default modification action plan (same structure as creation plans)
        """
        return {
            "sheet_info": {
                "sheet_name": sheet_name,
                "mode": "modify_unknown",
                "purpose": "Modification requested",
                "layout_type": "unknown"
            },
            "sections": [
                {
                    "section_name": "Entire Sheet",
                    "start_row": 1,
                    "current_state": "Unknown - no code or data available",
                    "changes": user_query[:500],
                    "data_items": [],
                    "formulas_to_modify": [],
                    "styling_changes": "To be determined"
                }
            ],
            "dependencies": [],
            "styling_requirements": {
                "changes_needed": "To be determined",
                "preserve": "All existing styling"
            },
            "modifications_summary": user_query[:200],
            "metadata": {
                "fallback": True,
                "reason": "Action plan generation failed or insufficient information",
                "has_existing_script": False
            }
        }
    
    def _get_default_sheet_action_plan(self, sheet_definition: Dict, sheet_index: int) -> Dict:
        """
        Get default action plan for a single sheet as fallback

        Args:
            sheet_definition: Sheet definition to base defaults on
            sheet_index: Index of the sheet

        Returns:
            Default action plan dictionary for this sheet
        """
        sheet_name = sheet_definition.get("name", f"Sheet{sheet_index+1}")
        
        return {
            "sheet_info": {
                "sheet_name": sheet_name,
                "sheet_index": sheet_index,
                "purpose": sheet_definition.get("description", "Data sheet"),
                "layout_type": sheet_definition.get("structure", {}).get("layout", "horizontal")
            },
            "sections": [
                {
                    "section_name": "Default Section",
                    "start_row": 1,
                    "row_count": 10,
                    "columns": {
                        "label_column": "A",
                        "data_columns": "auto"
                    },
                    "data_items": []
                }
            ],
            "dependencies": [],
            "styling_requirements": {
                "header_style": "Bold with blue background",
                "cell_styles": "Default formatting"
            },
            "metadata": {
                "fallback": True,
                "reason": "Sheet action plan generation failed, using default structure"
            }
        }

class CoordinatorAgent(BaseAgent):
    """Agent that coordinates all other agents"""
    
    def __init__(
        self,
        bedrock_client: BedrockClient,
        prompt_manager: PromptManager,
        session_manager=None,
    ):
        super().__init__(
            name="Coordinator",
            prompt_name="coordinator",
            bedrock_client=bedrock_client,
            prompt_manager=prompt_manager,
            temperature=0.7,
            session_manager=session_manager,
        )

class CoderAgent(BaseAgent):
    """Agent that generates Python code for Excel creation and editing"""

    def __init__(
        self,
        bedrock_client: BedrockClient,
        prompt_manager: PromptManager,
        session_manager=None,
    ):
        super().__init__(
            name="Coder",
            prompt_name="coder",
            bedrock_client=bedrock_client,
            prompt_manager=prompt_manager,
            temperature=0.3,  # Low temperature for precise code generation
            session_manager=session_manager,
        )

    def generate_sheet_code(
        self,
        user_query: str,
        assumptions: Dict,
        action_plan: Dict,
        workbook_path: str,
        is_first_sheet: bool = False
    ) -> str:
        """
        Generate Python code to create/edit a single Excel sheet

        Args:
            user_query: Original user request
            assumptions: Full assumptions configuration
            action_plan: Action plan for this specific sheet
            workbook_path: Path to the Excel workbook
            is_first_sheet: Whether this is the first sheet (create new workbook)

        Returns:
            Python code as string
        """
        self.logger.info(f"Generating code for sheet: {action_plan.get('sheet_info', {}).get('sheet_name', 'Unknown')}")

        try:
            system_prompt, user_template = self.prompt_manager.get_prompt_parts("coder_generate_sheet")
            user_message = user_template.format(
                user_query=user_query,
                assumptions=json.dumps(assumptions, indent=2),
                action_plan=json.dumps(action_plan, indent=2),
                workbook_path=workbook_path,
                is_first_sheet=str(is_first_sheet)
            )
            response = self.generate_response(user_message, system_prompt_override=system_prompt)

            # Clean response to remove markdown fences
            cleaned_code = clean_code_response(response)
            
            self.logger.info(f"Successfully generated code ({len(cleaned_code)} characters)")
            return cleaned_code

        except Exception as e:
            self.logger.error(f"Failed to generate sheet code", exc_info=e)
            raise

    def modify_sheet_code(
        self,
        user_query: str,
        existing_code: Optional[str],
        action_plan: Dict,
        assumptions: Dict,
        workbook_path: str,
        sheet_name: str
    ) -> str:
        """
        Modify existing sheet code based on user's modification request.
        
        If existing_code is None (e.g., modifying an external Excel file), generates
        fresh code that implements the modifications from the action plan.
        
        Args:
            user_query: User's modification request
            existing_code: Current code for the sheet (or None if not available)
            action_plan: Modification action plan
            assumptions: Full assumptions configuration
            workbook_path: Path to the Excel workbook
            sheet_name: Name of the sheet being modified
            
        Returns:
            Modified Python code as string
        """
        self.logger.info(f"Modifying code for sheet: {sheet_name}")
        
        # If no existing code, note that we're generating fresh modification code
        if not existing_code:
            self.logger.info("No existing code available; generating fresh code for modifications")
            existing_code = "# No existing code available - generating fresh modification script"
        
        try:
            system_prompt, user_template = self.prompt_manager.get_prompt_parts("coder_modify_sheet")
            user_message = user_template.format(
                user_query=user_query,
                sheet_name=sheet_name,
                existing_code=existing_code,
                action_plan=json.dumps(action_plan, indent=2),
                assumptions=json.dumps(assumptions, indent=2),
                workbook_path=workbook_path,
            )
            response = self.generate_response(user_message, system_prompt_override=system_prompt)
            
            # Clean response to remove markdown fences
            cleaned_code = clean_code_response(response)
            
            self.logger.info(f"Successfully modified code ({len(cleaned_code)} characters)")
            return cleaned_code
            
        except Exception as e:
            self.logger.error(f"Failed to modify sheet code", exc_info=e)
            # Fallback to original code
            return existing_code
    
    def edit_code(
        self,
        original_code: str,
        critic_feedback: str,
        execution_errors: Optional[List[str]] = None,
        action_plan: Dict = None
    ) -> str:
        """
        Edit existing code based on critic feedback and/or execution errors using structured edits

        Args:
            original_code: The original Python script
            critic_feedback: Feedback from critic agent
            execution_errors: Optional list of execution errors
            action_plan: Optional action plan for context

        Returns:
            Edited Python code as string
        """
        self.logger.info("Generating code edits based on feedback")

        try:
            system_prompt, user_template = self.prompt_manager.get_prompt_parts("coder_fix_code")
            
            error_summary = ""
            if execution_errors:
                error_summary = "EXECUTION ERRORS:\n" + "\n".join(f"- {err}" for err in execution_errors)

            user_message = user_template.format(
                original_code=original_code,
                critic_feedback=critic_feedback,
                execution_errors=error_summary,
                action_plan=json.dumps(action_plan, indent=2) if action_plan else "N/A"
            )
            response = self.generate_response(user_message, system_prompt_override=system_prompt)

            # Parse the JSON array of edits
            edits = self._parse_code_edits(response)
            
            if not edits:
                self.logger.warning("No edits parsed from LLM response, returning original code")
                return original_code
            
            self.logger.info(f"Parsed {len(edits)} code edits")
            
            # Apply edits to original code
            edited_code = self._apply_code_edits(original_code, edits)
            
            self.logger.info(f"Successfully edited code ({len(edited_code)} characters, {len(edits)} edits applied)")
            return edited_code

        except Exception as e:
            self.logger.error(f"Failed to edit code", exc_info=e)
            self.logger.warning("Falling back to original code due to edit failure")
            return original_code
    
    def _parse_code_edits(self, response: str) -> List[Dict]:
        """
        Parse JSON array of code edits from LLM response
        
        Args:
            response: Raw LLM response containing JSON array of edits
            
        Returns:
            List of edit dictionaries
        """
        try:
            # Use the parse_json utility function
            parsed = parse_json(response)
            
            if parsed and isinstance(parsed, list):
                self.logger.info(f"Successfully parsed {len(parsed)} edits from response")
                
                # Validate edit structure
                valid_edits = []
                for idx, edit in enumerate(parsed):
                    if not isinstance(edit, dict):
                        self.logger.warning(f"Edit {idx} is not a dictionary, skipping")
                        continue
                    
                    # Check required fields
                    if 'edit_type' not in edit or 'location' not in edit:
                        self.logger.warning(f"Edit {idx} missing required fields, skipping")
                        continue
                    
                    # Validate edit_type
                    if edit['edit_type'] not in ['replace', 'insert_after', 'insert_before', 'delete']:
                        self.logger.warning(f"Edit {idx} has invalid edit_type: {edit['edit_type']}, skipping")
                        continue
                    
                    valid_edits.append(edit)
                
                return valid_edits
            else:
                self.logger.error("Parsed response is not a list")
                return []
                
        except Exception as e:
            self.logger.error(f"Failed to parse code edits: {e}", exc_info=True)
            return []
    
    def _apply_code_edits(self, original_code: str, edits: List[Dict]) -> str:
        """
        Apply a list of edits to the original code

        Args:
            original_code: The original Python code
            edits: List of edit dictionaries

        Returns:
            Modified code with edits applied
        """
        import ast as _ast

        def _match_indent(new_code: str, reference_text: str) -> str:
            """
            If new_code lines have less indentation than the first non-empty line of
            reference_text, prepend the missing whitespace to each non-empty line.
            This prevents the LLM from inserting code at column 0 when it belongs
            inside an indented block.
            """
            ref_first = next((l for l in reference_text.split('\n') if l.strip()), '')
            ref_indent = len(ref_first) - len(ref_first.lstrip())
            if ref_indent == 0:
                return new_code
            new_lines = new_code.split('\n')
            new_first = next((l for l in new_lines if l.strip()), '')
            new_indent = len(new_first) - len(new_first.lstrip())
            if new_indent >= ref_indent:
                return new_code
            extra = ref_first[:ref_indent - new_indent]
            return '\n'.join(extra + l if l.strip() else l for l in new_lines)

        code = original_code
        applied_count = 0

        for idx, edit in enumerate(edits):
            try:
                edit_type = edit.get('edit_type')
                search_text = edit.get('search_text', '')
                new_code = edit.get('new_code', '')
                reason = edit.get('reason', 'No reason provided')

                self.logger.debug(f"Applying edit {idx+1}/{len(edits)}: {edit_type} - {reason}")

                if edit_type == 'replace':
                    if search_text in code:
                        code = code.replace(search_text, new_code, 1)  # Replace only first occurrence
                        applied_count += 1
                        self.logger.debug("  [OK] Replace successful")
                    else:
                        self.logger.warning("  [X] Replace failed: search_text not found in code")

                elif edit_type == 'insert_after':
                    if search_text in code:
                        parts = code.split(search_text, 1)
                        if len(parts) == 2:
                            adjusted = _match_indent(new_code, search_text)
                            code = parts[0] + search_text + '\n' + adjusted + parts[1]
                            applied_count += 1
                            self.logger.debug("  [OK] Insert after successful")
                    else:
                        self.logger.warning("  [X] Insert after failed: search_text not found")

                elif edit_type == 'insert_before':
                    if search_text in code:
                        parts = code.split(search_text, 1)
                        if len(parts) == 2:
                            adjusted = _match_indent(new_code, search_text)
                            code = parts[0] + adjusted + '\n' + search_text + parts[1]
                            applied_count += 1
                            self.logger.debug("  [OK] Insert before successful")
                    else:
                        self.logger.warning("  [X] Insert before failed: search_text not found")

                elif edit_type == 'delete':
                    if search_text in code:
                        code = code.replace(search_text, '', 1)  # Delete only first occurrence
                        applied_count += 1
                        self.logger.debug("  [OK] Delete successful")
                    else:
                        self.logger.warning("  [X] Delete failed: search_text not found")

            except Exception as e:
                self.logger.error(f"Failed to apply edit {idx+1}: {e}", exc_info=True)
                continue

        self.logger.info(f"Applied {applied_count}/{len(edits)} edits successfully")

        if applied_count == 0:
            self.logger.warning("No edits were applied - returning original code")
            return original_code

        # Validate syntax of the result. If edits introduced a syntax error, revert
        # to the original so the executor reports a meaningful error and the coder
        # can try again rather than silently producing an unrunnable script.
        try:
            _ast.parse(code)
        except SyntaxError as e:
            self.logger.error(
                f"Edits produced a syntax error (line {e.lineno}: {e.msg}); "
                "reverting to original code so the executor can surface the error."
            )
            return original_code

        return code

class ExecutorAgent(BaseAgent):
    """Agent that executes Python scripts for Excel generation"""

    # Whitelisted imports for safety
    ALLOWED_IMPORTS = {
        'openpyxl',
        'openpyxl.styles',
        'openpyxl.utils',
        'datetime',
        'math',
        're', 'pandas'
    }

    def __init__(
        self,
        bedrock_client: BedrockClient,
        prompt_manager: PromptManager,
        session_manager=None,
    ):
        super().__init__(
            name="Executor",
            prompt_name="executor",
            bedrock_client=bedrock_client,
            prompt_manager=prompt_manager,
            temperature=0.3,  # Low temperature for focused execution
            session_manager=session_manager,
        )
        self.execution_timeout = 300  # seconds (5 min; complex Excel scripts can be slow under MCP)
        self._session_manager = session_manager

    def _script_execution_log(
        self,
        script_path: str,
        workbook_path: str,
        success: bool,
        execution_time: float,
        return_code: Optional[int] = None,
        stdout: str = "",
        stderr: str = "",
        errors: Optional[List[str]] = None,
        script_content_preview: Optional[str] = None,
    ) -> None:
        """Write one entry to the session script_execution.log file."""
        try:
            sm = self._session_manager or get_session_manager()
            sm.log_script_execution(
                script_path=script_path,
                workbook_path=workbook_path,
                success=success,
                execution_time_seconds=execution_time,
                return_code=return_code,
                stdout=stdout or "",
                stderr=stderr or "",
                errors=errors,
                script_content_preview=script_content_preview,
            )
        except Exception as e:
            self.logger.warning(f"Could not write script execution log: {e}")

    def execute_script(self, script_path: str, workbook_path: str) -> Dict:
        """
        Execute a Python script to generate/modify Excel workbook

        Args:
            script_path: Path to Python script
            workbook_path: Path to workbook file

        Returns:
            Dictionary with execution results
        """
        self.logger.info(f"Executing script: {script_path}")

        # Read script content
        script_content = None
        script_preview = None
        try:
            script_content = Path(script_path).read_text(encoding='utf-8')
            script_preview = "\n".join(script_content.splitlines()[:35])
        except Exception as e:
            error_msg = f"Failed to read script: {str(e)}"
            self.logger.error(error_msg)
            self._script_execution_log(
                script_path, workbook_path,
                success=False, execution_time=0.0,
                errors=[error_msg], script_content_preview=None,
            )
            return {
                'success': False,
                'output_path': workbook_path,
                'errors': [error_msg],
                'execution_time': 0
            }

        # Validate script safety
        is_safe, safety_errors = self.validate_script_safety(script_content)
        if not is_safe:
            self.logger.error(f"Script failed safety validation: {safety_errors}")
            self._script_execution_log(
                script_path, workbook_path,
                success=False, execution_time=0.0,
                errors=safety_errors, script_content_preview=script_preview,
            )
            return {
                'success': False,
                'output_path': workbook_path,
                'errors': safety_errors,
                'execution_time': 0
            }

        # Execute script: resolve to absolute path to avoid cwd+relative path doubling
        start_time = time.time()
        script_path_resolved = Path(script_path).resolve()
        script_dir = script_path_resolved.parent
        try:
            result = subprocess.run(
                # IMPORTANT (Windows/MCP): use the same interpreter running this server,
                # not a PATH-resolved "python" shim/alias which can hang under Cursor/MCP.
                [sys.executable, str(script_path_resolved), workbook_path],
                capture_output=True,
                text=True,
                timeout=self.execution_timeout,
                cwd=str(script_dir),
                # Prevent child waiting for stdin (can deadlock in some environments).
                stdin=subprocess.DEVNULL,
            )

            execution_time = time.time() - start_time

            if result.returncode == 0:
                self.logger.info(f"Script executed successfully in {execution_time:.2f}s")
                self._script_execution_log(
                    script_path, workbook_path,
                    success=True, execution_time=execution_time,
                    return_code=result.returncode,
                    stdout=result.stdout or "", stderr=result.stderr or "",
                    errors=[], script_content_preview=script_preview,
                )
                return {
                    'success': True,
                    'output_path': workbook_path,
                    'stdout': result.stdout,
                    'errors': [],
                    'execution_time': execution_time
                }
            else:
                error_msg = f"Script failed with return code {result.returncode}"
                self.logger.error(f"{error_msg}\nStderr: {result.stderr}")
                err_list = [error_msg, result.stderr] if result.stderr else [error_msg]
                self._script_execution_log(
                    script_path, workbook_path,
                    success=False, execution_time=execution_time,
                    return_code=result.returncode,
                    stdout=result.stdout or "", stderr=result.stderr or "",
                    errors=err_list, script_content_preview=script_preview,
                )
                return {
                    'success': False,
                    'output_path': workbook_path,
                    'stdout': result.stdout,
                    'stderr': result.stderr,
                    'errors': [error_msg, result.stderr],
                    'execution_time': execution_time
                }

        except subprocess.TimeoutExpired:
            execution_time = time.time() - start_time
            error_msg = f"Script execution timeout after {self.execution_timeout}s"
            self.logger.error(error_msg)
            self._script_execution_log(
                script_path, workbook_path,
                success=False, execution_time=execution_time,
                errors=[error_msg], script_content_preview=script_preview,
            )
            return {
                'success': False,
                'output_path': workbook_path,
                'errors': [error_msg],
                'execution_time': execution_time
            }
        except Exception as e:
            execution_time = time.time() - start_time
            error_msg = f"Script execution failed: {str(e)}"
            self.logger.error(error_msg, exc_info=e)
            self._script_execution_log(
                script_path, workbook_path,
                success=False, execution_time=execution_time,
                errors=[error_msg], script_content_preview=script_preview,
            )
            return {
                'success': False,
                'output_path': workbook_path,
                'errors': [error_msg],
                'execution_time': execution_time
            }

    def validate_script_safety(self, script_content: str) -> Tuple[bool, List[str]]:
        """
        Validate that script is safe to execute

        Args:
            script_content: Python script content

        Returns:
            Tuple of (is_safe, list_of_errors)
        """
        errors = []

        # 1. Check syntax
        try:
            ast.parse(script_content)
        except SyntaxError as e:
            errors.append(f"Syntax error at line {e.lineno}: {e.msg}")
            return False, errors

        is_safe = len(errors) == 0
        return is_safe, errors

