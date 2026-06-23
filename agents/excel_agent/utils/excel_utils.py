"""
Excel utilities for file operations and analysis
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional
import pandas as pd
from pathlib import Path
from .logging_utils import get_logger


class ExcelManager:
    """Manager for Excel file operations"""

    def __init__(self, session_manager=None):
        """Initialize Excel manager"""
        self.logger = get_logger("ExcelManager", session_manager=session_manager)
        self.current_workbook = None
        self.current_file = None
    
    def create_workbook(self) -> openpyxl.Workbook:
        """
        Create a new Excel workbook
        
        Returns:
            openpyxl Workbook object
        """
        try:
            wb = openpyxl.Workbook()
            self.current_workbook = wb
            self.logger.info("Created new workbook")
            return wb
        except Exception as e:
            self.logger.error("Failed to create workbook", exc_info=e)
            raise
    
    def load_workbook(self, filepath: str) -> openpyxl.Workbook:
        """
        Load existing Excel workbook
        
        Args:
            filepath: Path to Excel file
            
        Returns:
            openpyxl Workbook object
        """
        try:
            wb = openpyxl.load_workbook(filepath)
            self.current_workbook = wb
            self.current_file = filepath
            self.logger.info(f"Loaded workbook from {filepath}")
            return wb
        except Exception as e:
            self.logger.error(
                f"Failed to load workbook from {filepath}",
                exc_info=e
            )
            raise
    
    def save_workbook(self, filepath: str, workbook: Optional[openpyxl.Workbook] = None):
        """
        Save workbook to file
        
        Args:
            filepath: Output file path
            workbook: Workbook to save (uses current if None)
        """
        try:
            wb = workbook or self.current_workbook
            if not wb:
                raise ValueError("No workbook to save")
            
            wb.save(filepath)
            self.current_file = filepath
            self.logger.info(f"Saved workbook to {filepath}")
        except Exception as e:
            self.logger.error(
                f"Failed to save workbook to {filepath}",
                exc_info=e
            )
            raise
    
    def analyze_workbook(self, workbook: Optional[openpyxl.Workbook] = None) -> Dict:
        """
        Analyze workbook structure and content
        
        Args:
            workbook: Workbook to analyze (uses current if None)
            
        Returns:
            Dictionary with analysis results
        """
        try:
            wb = workbook or self.current_workbook
            if not wb:
                raise ValueError("No workbook to analyze")
            
            analysis = {
                'sheet_count': len(wb.sheetnames),
                'sheets': {},
                'total_formulas': 0,
                'total_cells_with_data': 0
            }
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                sheet_info = {
                    'name': sheet_name,
                    'max_row': sheet.max_row,
                    'max_column': sheet.max_column,
                    'formulas': [],
                    'data_cells': 0
                }
                
                # Analyze cells
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            sheet_info['data_cells'] += 1
                            
                            # Check if it's a formula
                            if isinstance(cell.value, str) and cell.value.startswith('='):
                                sheet_info['formulas'].append({
                                    'cell': cell.coordinate,
                                    'formula': cell.value
                                })
                                analysis['total_formulas'] += 1
                
                analysis['sheets'][sheet_name] = sheet_info
                analysis['total_cells_with_data'] += sheet_info['data_cells']
            
            self.logger.info(f"Analyzed workbook: {analysis['sheet_count']} sheets, "
                           f"{analysis['total_formulas']} formulas")
            return analysis
            
        except Exception as e:
            self.logger.error("Failed to analyze workbook", exc_info=e)
            raise
    
    def get_cell_value(self, sheet_name: str, cell: str, 
                       workbook: Optional[openpyxl.Workbook] = None) -> Any:
        """
        Get value from specific cell
        
        Args:
            sheet_name: Name of sheet
            cell: Cell reference (e.g., 'A1')
            workbook: Workbook to use (uses current if None)
            
        Returns:
            Cell value
        """
        try:
            wb = workbook or self.current_workbook
            if not wb:
                raise ValueError("No workbook available")
            
            value = wb[sheet_name][cell].value
            self.logger.debug(f"Retrieved {sheet_name}!{cell} = {value}")
            return value
            
        except Exception as e:
            self.logger.error(
                f"Failed to get value from {sheet_name}!{cell}",
                exc_info=e
            )
            raise
    
    def set_cell_value(self, sheet_name: str, cell: str, value: Any,
                       workbook: Optional[openpyxl.Workbook] = None):
        """
        Set value in specific cell
        
        Args:
            sheet_name: Name of sheet
            cell: Cell reference (e.g., 'A1')
            value: Value to set
            workbook: Workbook to use (uses current if None)
        """
        try:
            wb = workbook or self.current_workbook
            if not wb:
                raise ValueError("No workbook available")
            
            wb[sheet_name][cell].value = value
            self.logger.debug(f"Set {sheet_name}!{cell} = {value}")
            
        except Exception as e:
            self.logger.error(
                f"Failed to set value in {sheet_name}!{cell}",
                exc_info=e
            )
            raise
    
    def apply_formatting(self, sheet_name: str, cell: str, 
                        font: Optional[Dict] = None,
                        fill: Optional[Dict] = None,
                        alignment: Optional[Dict] = None,
                        workbook: Optional[openpyxl.Workbook] = None):
        """
        Apply formatting to a cell
        
        Args:
            sheet_name: Name of sheet
            cell: Cell reference
            font: Font properties (bold, color, etc.)
            fill: Fill properties (background color)
            alignment: Alignment properties
            workbook: Workbook to use (uses current if None)
        """
        try:
            wb = workbook or self.current_workbook
            if not wb:
                raise ValueError("No workbook available")
            
            cell_obj = wb[sheet_name][cell]
            
            if font:
                cell_obj.font = Font(**font)
            if fill:
                cell_obj.fill = PatternFill(**fill)
            if alignment:
                cell_obj.alignment = Alignment(**alignment)
            
            self.logger.debug(f"Applied formatting to {sheet_name}!{cell}")
            
        except Exception as e:
            self.logger.error(
                f"Failed to apply formatting to {sheet_name}!{cell}",
                exc_info=e
            )
            raise
    
    def detect_errors(self, workbook: Optional[openpyxl.Workbook] = None) -> Dict:
        """
        Detect formula errors in workbook.
        Checks both formula cells (for error literals) and loads with data_only
        to get calculated values (#REF!, #DIV/0! etc. from Excel).
        """
        try:
            wb = workbook or self.current_workbook
            if not wb:
                raise ValueError("No workbook available")

            error_types = ['#REF!', '#DIV/0!', '#VALUE!', '#N/A', '#NAME?', '#NULL!', '#NUM!']
            errors = {error: [] for error in error_types}

            # Check formula cells in current wb (data_only=False)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value in error_types:
                            errors[cell.value].append(f"{sheet_name}!{cell.coordinate}")

            # Also load with data_only=True to get calculated errors (if file was opened in Excel)
            try:
                fp = self.current_file
                if fp and Path(fp).exists():
                    wb_calc = openpyxl.load_workbook(fp, data_only=True)
                    for sheet_name in wb_calc.sheetnames:
                        sheet = wb_calc[sheet_name]
                        for row in sheet.iter_rows():
                            for cell in row:
                                val = cell.value
                                if isinstance(val, str) and val.strip() in error_types:
                                    loc = f"{sheet_name}!{cell.coordinate}"
                                    if loc not in errors.get(val.strip(), []):
                                        errors[val.strip()].append(loc)
            except Exception:
                pass

            total_errors = sum(len(locations) for locations in errors.values())
            if total_errors > 0:
                self.logger.warning(f"Found {total_errors} formula errors")
            else:
                self.logger.info("No formula errors detected")
            return errors
        except Exception as e:
            self.logger.error("Failed to detect errors", exc_info=e)
            raise

    def extract_sheet_content(
        self,
        workbook: Optional[openpyxl.Workbook],
        sheet_name: str,
        max_rows: int = 80,
        max_cols: int = 30,
    ) -> Dict:
        """
        Extract actual sheet content for critic review: cells, formulas, values.
        Returns structure the critic can use to find wrong refs, empty cells, errors.
        """
        wb = workbook or self.current_workbook
        if not wb:
            raise ValueError("No workbook available")
        if sheet_name not in wb.sheetnames:
            return {"error": f"Sheet '{sheet_name}' not found", "cells": [], "formulas": [], "sheet_names": wb.sheetnames}

        sheet = wb[sheet_name]
        cells = []
        formulas = []
        max_r = min(sheet.max_row, max_rows)
        max_c = min(sheet.max_column, max_cols)

        for row in sheet.iter_rows(min_row=1, max_row=max_r, max_col=max_c):
            for cell in row:
                val = cell.value
                if val is None:
                    continue
                coord = cell.coordinate
                is_formula = isinstance(val, str) and val.startswith("=")
                entry = {
                    "coord": coord,
                    "formula": val if is_formula else None,
                    "value": str(val)[:80] if not is_formula else "(formula)",
                }
                cells.append(entry)
                if is_formula:
                    formulas.append({"cell": coord, "formula": val})

        # Load data_only for calculated values (including errors)
        values_resolved = {}
        try:
            fp = self.current_file
            if fp and Path(fp).exists():
                wb_calc = openpyxl.load_workbook(fp, data_only=True)
                if sheet_name in wb_calc.sheetnames:
                    ws = wb_calc[sheet_name]
                    for row in ws.iter_rows(min_row=1, max_row=max_r, max_col=max_c):
                        for cell in row:
                            if cell.value is not None:
                                values_resolved[cell.coordinate] = str(cell.value)[:80]
        except Exception:
            pass

        # Build grid summary (rows 1-20, cols A-J for readability; plus all formulas)
        grid_lines = []
        for r in range(1, min(21, max_r + 1)):
            row_vals = []
            for c in range(1, min(11, max_c + 1)):
                col_letter = get_column_letter(c)
                coord = f"{col_letter}{r}"
                cell = sheet[coord]
                v = cell.value
                if v is None:
                    row_vals.append("")
                elif isinstance(v, str) and v.startswith("="):
                    row_vals.append(f"[F]={v[:50]}" + ("..." if len(v) > 50 else ""))
                else:
                    row_vals.append(str(v)[:20])
            grid_lines.append(f"Row {r}: " + " | ".join(row_vals))

        return {
            "sheet_name": sheet_name,
            "sheet_names": list(wb.sheetnames),
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "grid_sample": "\n".join(grid_lines),
            "formulas": formulas[:200],
            "values_resolved": values_resolved,
            "cell_count": len(cells),
        }
