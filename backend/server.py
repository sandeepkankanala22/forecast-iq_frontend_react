"""
forecast_server.py
------------------
FastAPI backend for the React frontend (frontend/dist).

Endpoints
---------
GET  /              → serves the React SPA
POST /api/chat      → Bedrock-powered chat assistant
POST /api/research  → Bedrock-powered epidemiology / market research

Run
---
    uvicorn forecast_server:app --reload --port 8000

Environment variables (.env or shell)
--------------------------------------
    MODEL_ID=us.anthropic.claude-sonnet-4-5-20250929-v1:0  # optional
    AWS_REGION=us-east-1                                    # optional
    S3_BUCKET=<bucket-name>                                 # optional, enables S3 for outputs/logs
    S3_PREFIX=forecast-agent                                 # optional, S3 key prefix
    S3_REGION=us-east-1                                     # optional, defaults to AWS_REGION
"""

import multiprocessing
multiprocessing.freeze_support()

import asyncio
import json
import logging
import math
import os
import re
import shutil
import sys
import tempfile
import time
import uuid
from datetime import datetime,date
from pathlib import Path
from typing import Dict, List, Optional

# Project root is two levels up from backend/server.py
_project_root = Path(__file__).resolve().parent.parent
_agent_dir = _project_root / "agents" / "excel_agent"
_DATA_OUTPUT_DIR = _project_root / "data" / "output"  # legacy fallback
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))

import boto3
from botocore.config import Config
from botocore.exceptions import ClientError
from dotenv import load_dotenv
load_dotenv(_project_root / ".env")

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

try:
    from backend.services.s3_storage import (
        s3_enabled,
        upload_file,
        upload_json,
        get_presigned_download_url,
        download_to_temp,
        s3_output_key,
        s3_logs_key,
    )
except ImportError:

    def s3_enabled():
        return False

    def _noop(*a, **k):
        return None

    upload_file = upload_json = get_presigned_download_url = download_to_temp = _noop
    s3_output_key = s3_logs_key = lambda *p: "/".join(p) if p else ""

try:
    from urllib3.exceptions import ReadTimeoutError
except ImportError:
    ReadTimeoutError = TimeoutError


# ---------------------------------------------------------------------------
# Bedrock logger stub (for BedrockClient)
# ---------------------------------------------------------------------------
def get_logger(name: str) -> logging.Logger:
    logger = logging.getLogger(name)
    if not logger.handlers:
        h = logging.StreamHandler()
        h.setFormatter(logging.Formatter("%(asctime)s %(name)s %(levelname)s %(message)s"))
        logger.addHandler(h)
        logger.setLevel(logging.INFO)
    return logger


class _BedrockLoggerAdapter:
    """Adapter to provide BedrockClient's expected logger interface."""
    def __init__(self, name: str):
        self._log = get_logger(name)

    def info(self, msg: str, **kwargs): self._log.info(msg, **kwargs)
    def warning(self, msg: str, **kwargs): self._log.warning(msg, **kwargs)
    def error(self, msg: str, **kwargs): self._log.error(msg, **kwargs)
    def critical(self, msg: str, **kwargs): self._log.critical(msg, **kwargs)

    def log_function_call(self, name: str, data: dict):
        self._log.info("bedrock %s %s", name, json.dumps(data, default=str))

    def log_bedrock_call(self, call_id, model, input_messages, system_prompt, output, time_elapsed, tokens_used, metadata=None):
        self._log.info("bedrock call_id=%s model=%s time=%.2fs tokens=%s", call_id, model, time_elapsed, tokens_used)

    def log_error_with_context(self, error, context: dict):
        self._log.error("bedrock error: %s context=%s", error, json.dumps(context, default=str))


def get_bedrock_logger(name: str) -> _BedrockLoggerAdapter:
    return _BedrockLoggerAdapter(name)


# ---------------------------------------------------------------------------
# Bedrock client
# ---------------------------------------------------------------------------
class BedrockClient:
    """Wrapper for AWS Bedrock API calls"""

    def __init__(
        self,
        region: str = "us-east-1",
        read_timeout: int = 600,
        model_id: Optional[str] = None,
    ):
        default_id = "us.anthropic.claude-sonnet-4-5-20250929-v1:0"
        self.logger = get_bedrock_logger("BedrockClient")
        self.region = region
        self.model_id = model_id if model_id else os.getenv("MODEL_ID", default_id)
        self.read_timeout = read_timeout

        config = Config(
            read_timeout=read_timeout,
            connect_timeout=10,
            retries={"max_attempts": 3, "mode": "adaptive"},
        )
        self.client = boto3.client(
            service_name="bedrock-runtime",
            region_name=region,
            config=config,
        )
        self.logger.info(f"Bedrock client initialized in region {region} with {read_timeout}s read timeout")

    def invoke_model(
        self,
        messages: List[Dict],
        system_prompt: Optional[str] = None,
        max_tokens: int = 4096,
        temperature: Optional[float] = 0.7,
        top_p: Optional[float] = None,
        metadata: Optional[Dict] = None,
    ) -> Dict:
        call_id = str(uuid.uuid4())[:8]
        start_time = time.time()

        try:
            body = {
                "anthropic_version": "bedrock-2023-05-31",
                "messages": messages,
                "max_tokens": max_tokens,
            }
            if top_p is not None:
                body["top_p"] = top_p
            elif temperature is not None:
                body["temperature"] = temperature

            if system_prompt:
                body["system"] = system_prompt

            prompt_length = sum(len(str(m.get("content", ""))) for m in messages)
            self.logger.log_function_call(
                "invoke_model",
                {"call_id": call_id, "model": self.model_id, "messages_count": len(messages), "prompt_length": prompt_length, "max_tokens": max_tokens},
            )

            response = self.client.invoke_model(modelId=self.model_id, body=json.dumps(body))
            time_elapsed = time.time() - start_time

            response_body = json.loads(response["body"].read())

            content = ""
            if response_body.get("content"):
                for block in response_body["content"]:
                    if block.get("type") == "text":
                        content += block.get("text", "")

            tokens_used = response_body.get("usage", {})

            self.logger.log_bedrock_call(
                call_id=call_id,
                model=self.model_id,
                input_messages=messages,
                system_prompt=system_prompt,
                output=content,
                time_elapsed=time_elapsed,
                tokens_used=tokens_used,
                metadata=metadata,
            )

            return {
                "content": content,
                "stop_reason": response_body.get("stop_reason"),
                "usage": tokens_used,
                "model": self.model_id,
                "call_id": call_id,
                "time_elapsed": time_elapsed,
            }

        except ClientError as e:
            time_elapsed = time.time() - start_time
            error_code = e.response.get("Error", {}).get("Code", "Unknown")
            error_message = e.response.get("Error", {}).get("Message", str(e))
            self.logger.log_error_with_context(
                error=e,
                context={
                    "call_id": call_id,
                    "model": self.model_id,
                    "error_code": error_code,
                    "error_message": error_message,
                    "messages_count": len(messages),
                    "time_elapsed": time_elapsed,
                },
            )
            raise
        except Exception as e:
            time_elapsed = time.time() - start_time
            self.logger.log_error_with_context(
                error=e,
                context={
                    "call_id": call_id,
                    "model": self.model_id,
                    "messages_count": len(messages),
                    "function": "invoke_model",
                    "time_elapsed": time_elapsed,
                },
            )
            raise

    def invoke_with_retry(
        self,
        messages: List[Dict],
        system_prompt: Optional[str] = None,
        max_retries: int = 3,
        **kwargs,
    ) -> Dict:
        last_exception = None

        for attempt in range(max_retries):
            try:
                metadata = kwargs.get("metadata", {})
                metadata["retry_attempt"] = attempt + 1
                kwargs["metadata"] = metadata

                return self.invoke_model(
                    messages=messages,
                    system_prompt=system_prompt,
                    **kwargs,
                )
            except ClientError as e:
                last_exception = e
                error_code = e.response.get("Error", {}).get("Code", "")

                if error_code in ["ValidationException", "AccessDeniedException"]:
                    raise

                wait_time = 2**attempt
                self.logger.warning(f"Bedrock call failed (attempt {attempt + 1}/{max_retries}). Retrying in {wait_time}s. Error: {error_code}")
                time.sleep(wait_time)
            except ReadTimeoutError as e:
                last_exception = e
                self.logger.warning(
                    f"Read timeout on attempt {attempt + 1}/{max_retries}. LLM response took too long (>{self.read_timeout}s). "
                    "Consider reducing input size or increasing timeout."
                )
                if attempt == max_retries - 1:
                    self.logger.error("All retry attempts exhausted due to read timeouts.")
                    raise
                wait_time = 5
                time.sleep(wait_time)
            except Exception as e:
                last_exception = e
                self.logger.error(f"Unexpected error on attempt {attempt + 1}/{max_retries}", exc_info=e)
                if attempt == max_retries - 1:
                    raise
                time.sleep(2**attempt)

        self.logger.critical("All retries failed", exc_info=last_exception)
        raise last_exception


_bedrock = BedrockClient(
    region=os.getenv("AWS_REGION", "us-east-1"),
    read_timeout=600,
)

# ---------------------------------------------------------------------------
# App & CORS
# ---------------------------------------------------------------------------
app = FastAPI(title="Forecast AI Backend", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# Request / Response models
# ---------------------------------------------------------------------------

class Message(BaseModel):
    role: str     # "user" | "assistant"
    content: str


class ChatRequest(BaseModel):
    messages: list[Message]
    chat_step: int = 0
    form_state: dict = {}
    workflow_stage: str = "product_info"   # product_info | parameter_selection | assumptions | forecast_engine | results


class ChatResponse(BaseModel):
    bot_message: str
    field_updates: dict = {}
    quick_replies: list[str] = []
    # one of: "", "show_parameter_selection", "generate_assumptions",
    #         "calculate_forecast", "proceed_results", "start_over"
    action: str = ""


class ResearchRequest(BaseModel):
    indication: str
    country: str
    class_moa: str


# ---------------------------------------------------------------------------
# Serve the React frontend (SPA)
# ---------------------------------------------------------------------------

_FRONTEND_DIST = _project_root / "frontend" / "dist"
_FRONTEND_LEGACY = _project_root / "frontend-legacy"


def _spa_index() -> FileResponse:
    """Return the built React index.html for client-side routes."""
    index = _FRONTEND_DIST / "index.html"
    if index.is_file():
        return FileResponse(index, media_type="text/html")
    # Fallback to legacy HTML during local dev without a frontend build
    legacy = _FRONTEND_LEGACY / "Forecast_Integrated_Chat.html"
    if legacy.is_file():
        return FileResponse(legacy, media_type="text/html")
    raise HTTPException(
        status_code=404,
        detail="Frontend not found. Run: cd frontend && npm install && npm run build",
    )


def _static_asset(*candidates: Path, media_type: str) -> FileResponse:
    for path in candidates:
        if path.is_file():
            return FileResponse(path, media_type=media_type)
    raise HTTPException(status_code=404, detail="Asset not found")


@app.get("/whitebglogo.svg")
async def serve_logo():
    return _static_asset(
        _FRONTEND_DIST / "whitebglogo.svg",
        _project_root / "whitebglogo.svg",
        media_type="image/svg+xml",
    )


@app.get("/image.webp")
async def serve_profile_image():
    return _static_asset(
        _FRONTEND_DIST / "image.webp",
        _project_root / "image.webp",
        media_type="image/webp",
    )


@app.get("/")
async def serve_frontend():
    return _spa_index()


# ---------------------------------------------------------------------------
# Chat endpoint
# ---------------------------------------------------------------------------

CHAT_SYSTEM = """\
You are an expert AI Forecast Assistant embedded inside a commercial pharmaceutical forecasting tool.
You help users build revenue forecasts for pharmaceutical / biotech assets.

═══ TOOL WORKFLOW — 5 stages ═══════════════════════════════════════════════════════
Stage 1  product_info          – Collect 6 product fields (see below)
Stage 2  parameter_selection   – User picks which epidemiology / market parameters to include
Stage 3  assumptions           – AI-generated assumptions (prevalence, diagnosis rate, etc.) shown in table
Stage 4  forecast_engine       – User sets forecast horizon, pricing, curve settings
Stage 5  results               – Charts and revenue tables displayed
═════════════════════════════════════════════════════════════════════════════════════

You will receive:
  • "workflow_stage" – which stage the user is currently on
  • "form_state"     – current values of the 6 product fields (already filled ones)
  • "chat_step"      – how many of the 6 fields have been filled (0–6)
  • "Next field to collect" – the first EMPTY field name

───── Stage 1 fields ─────────────────────────────────────────────────────────────
  1. country       – MUST be one of exactly: United States, Germany, United Kingdom,
                     France, Japan, China, Canada, Italy, Spain
  2. productName   – drug / compound name (any reasonable string)
  3. classMoa      – drug class or mechanism of action (e.g. "Monoclonal Antibody",
                     "SGLT2 Inhibitor", "PD-1 Inhibitor", "BTK Inhibitor", "JAK Inhibitor")
  4. indication    – therapeutic area (e.g. Rheumatoid Arthritis, Oncology,
                     Multiple Sclerosis, Type 2 Diabetes, Heart Failure)
  5. launchYear    – planned launch year (integer 2024–2040)
  6. peakYear      – forecast end year (MUST be > launchYear, integer 2025–2045)

═══ MULTI-FIELD EXTRACTION (CRITICAL) ══════════════════════════════════════════════
When a user sends a single message that contains MULTIPLE field values — for example:
  "I want to build a forecast for Keytruda, a PD-1 inhibitor, in the US for NSCLC,
   launching in 2026 with forecast end year 2031"
— you MUST extract ALL provided fields at once into field_updates:
  { "country": "United States", "productName": "Keytruda",
    "classMoa": "PD-1 Inhibitor", "indication": "NSCLC",
    "launchYear": "2026", "peakYear": "2031" }

After extracting, identify which fields are STILL missing and ask for the next one.

═══ VALIDATION & NONSENSICAL INPUT HANDLING ════════════════════════════════════════
Before populating field_updates, validate each extracted value:
  • country       : Must match one of the 9 allowed countries (case-insensitive).
                    If user says "UK" treat as "United Kingdom", "US"→"United States", etc.
                    If unrecognisable, do NOT add to field_updates; ask for correction.
  • launchYear    : Must be 2024–2040. If out of range or clearly wrong (e.g. 1990, 2120),
                    do NOT add; ask for a realistic year.
  • peakYear (Forecast - End Year) : Must be 2025–2045 AND > launchYear. Validate after launchYear is known. Always refer to this field as "Forecast - End Year" when asking the user.
                    IMPORTANT: If the user sends ONLY a 4-digit year (e.g. "2031", "2032") and
                    peakYear is the next missing field, treat it directly as peakYear — do NOT
                    say you didn't understand it. Similarly for launchYear if that is missing.
  • productName   : Any non-empty string is valid. However if it looks like random noise
                    (e.g. "asdfgh", "123456", "the sky is blue"), flag it:
                    ask "Just to confirm — is **[value]** really the product name, or did
                    you mean something else?"
  • classMoa      : Any non-empty string. If it looks completely unrelated to medicine,
                    ask for clarification.
  • indication    : Any non-empty string. If it looks completely unrelated to a disease,
                    ask for clarification.

If the user sends a message that is COMPLETELY off-topic or nonsensical (greeting aside),
do NOT add anything to field_updates. Instead explain what information is needed and
ask for the next missing field.

═══ YOUR BEHAVIOUR BY STAGE ════════════════════════════════════════════════════════

▸ Stage 1 (product_info):
  - SCAN the entire message for all 6 field values and extract all you find.
  - DO NOT extract a field value if the user is clearly asking a question about it
    (contains "?", starts with question words like what/how/which/can/could/tell/suggest).
  - After extraction, ask for the FIRST still-missing field. Never ask for an already-filled field.
  - When all 6 fields are captured tell the user all 6 are done and set
    action = "show_parameter_selection".
  - For domain questions ("what is a good forecast - end year?"), give expert answer then
    redirect to the next missing field.
  - quick_replies: context-specific per next missing field:
      • country    → the 9 allowed country names
      • productName → [] (empty — product names are unique, do not suggest any)
      • classMoa   → ["Monoclonal Antibody", "SGLT2 Inhibitor", "PD-1 Inhibitor", "JAK Inhibitor", "BTK Inhibitor"]
      • indication → ["Oncology", "Type 2 Diabetes", "Rheumatoid Arthritis", "Heart Failure", "Multiple Sclerosis"]
      • launchYear → ["2025", "2026", "2027", "2028", "2030"]
      • peakYear   → ["2030", "2031", "2032", "2033", "2035"]

▸ Stage 2 (parameter_selection):
  - User is choosing which parameters to include (prevalence vs incidence, etc.).
  - Answer questions about what each parameter means.
  - quick_replies: ["Generate Assumptions", "What is prevalence?", "What is class share?"]

▸ Stage 3 (assumptions):
  - Assumptions table is visible. User can review and edit values.
  - Answer specific questions about any assumption.
  - Suggest calculating the forecast when ready.
  - quick_replies: ["Calculate Forecast", "What does diagnosis rate mean?", "Edit assumptions"]

▸ Stage 4 (forecast_engine):
  - Guide user to run the forecast calculation.
  - quick_replies: ["Calculate Forecast", "Change Forecast - End Year", "What is S-curve adoption?"]

▸ Stage 5 (results):
  - User can view charts and export.
  - quick_replies: ["Export to CSV", "Start new forecast", "Explain the chart"]

═══ DOMAIN KNOWLEDGE ════════════════════════════════════════════════════════════════
You have deep expertise in:
  • Epidemiology (prevalence, incidence, diagnosis rates, treatment rates)
  • Pharmaceutical market access and pricing (US, EU5, Japan)
  • Drug classes & mechanisms of action
  • Commercial forecasting methodology (patient funnel, S-curve adoption, market share)
  • Real-world data sources (IDF, WHO, ACR, GLOBOCAN, MSIF Atlas)

Never be vague. Never say "I'm just an AI" or "I cannot answer that".

═══ OUTPUT FORMAT ════════════════════════════════════════════════════════════════════
Respond ONLY with a JSON object (no prose outside it):
{
  "bot_message":    "<string – your reply, **bold** key terms, mention each field you just filled>",
  "field_updates":  { "<fieldName>": "<value>" },
  "quick_replies":  ["<chip1>", "<chip2>"],
  "action":         "<string>"
    // action choices:
    //   "show_parameter_selection"  → ALL 6 fields are now filled (auto-proceed)
    //   "generate_assumptions"      → user explicitly says "generate assumptions"
    //   "calculate_forecast"        → user explicitly says "calculate" / "run forecast"
    //   "proceed_results"           → user explicitly says "results" / "show charts"
    //   "start_over"                → user explicitly says "start over" / "restart"
}

Critical rules:
  • field_updates: extract ALL fields mentioned; leave empty {} if none found or input invalid.
  • In bot_message: confirm each field you just filled, then ask for the next missing field.
  • NEVER ask for a field that you are including in field_updates — if it's in field_updates
    it is already captured; move on to the next truly missing field.
  • If field_updates fills ALL remaining missing fields, do NOT ask for anything —
    instead confirm all 6 are collected and set action="show_parameter_selection".
  • quick_replies: 2–5 short context-specific options (≤ 30 chars). Never "OK" or "Continue".
  • action: set "show_parameter_selection" automatically the moment all 6 fields are filled.
  • bot_message: 2–4 sentences normally; may be longer for domain questions.
"""


@app.post("/api/chat", response_model=ChatResponse)
async def chat(req: ChatRequest):
    # Build system message with current context injected
    system_content = CHAT_SYSTEM

    # Describe what fields are filled vs empty
    form = req.form_state
    filled   = [k for k, v in form.items() if v]
    missing  = [k for k, v in form.items() if not v]
    context_lines = [
        f"\n\n══ CURRENT CONTEXT ══",
        f"workflow_stage : {req.workflow_stage}",
        f"fields_filled  : {len(filled)}/6",
        f"fields_filled_list : {filled if filled else '(none yet)'}",
        f"fields_missing_list: {missing if missing else '(all filled!)'}",
    ]
    if form:
        context_lines.append("form_state (current values):")
        for k, v in form.items():
            status = "✓ FILLED" if v else "✗ EMPTY — collect this"
            context_lines.append(f"  {k}: {v!r}  [{status}]")
    if missing:
        context_lines.append(f"\nNext field to collect: {missing[0]}")
        context_lines.append(f"All remaining missing fields: {', '.join(missing)}")
        context_lines.append(
            "\nIMPORTANT: Scan the user's message for ANY of the missing fields. "
            "Extract all you can, then ask only for whatever is still missing."
        )
    else:
        context_lines.append("\nALL 6 FIELDS ARE FILLED — set action=show_parameter_selection immediately.")
    system_content += "\n".join(context_lines)

    # Bedrock: system goes in system_prompt; messages = user/assistant only
    bedrock_messages = []
    for m in req.messages[-20:]:
        if m.role in ("user", "assistant"):
            bedrock_messages.append({"role": m.role, "content": m.content})

    resp = await asyncio.to_thread(
        _bedrock.invoke_with_retry,
        messages=bedrock_messages,
        system_prompt=system_content + "\n\nRespond ONLY with a valid JSON object. No prose outside the JSON.",
        temperature=0.2,
        max_tokens=600,
    )

    raw = resp["content"]
    cleaned = re.sub(r"^```(?:json)?\s*", "", raw.strip(), flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*```$", "", cleaned.strip())
    m = re.search(r"\{[\s\S]*\}", cleaned)
    if m:
        cleaned = m.group(0)
    try:
        data = json.loads(cleaned)
    except json.JSONDecodeError:
        data = {}

    # ── Server-side safety: strip empty field_updates values so the client
    #    never overwrites already-filled fields with empty strings.
    raw_updates = data.get("field_updates") or {}
    field_updates = {k: v for k, v in raw_updates.items() if v is not None and str(v).strip() != ""}

    # ── Compute which fields are STILL missing after applying field_updates ──
    #    This prevents the LLM from asking for a field it just extracted.
    all_field_keys = ["country", "productName", "classMoa", "indication", "launchYear", "peakYear"]
    merged_form = {**form, **field_updates}   # form_state + what was just extracted
    still_missing = [k for k in all_field_keys if not merged_form.get(k)]

    # ── Server-side safety: validate the returned action against reality.
    #    The LLM sometimes hallucinates destructive actions for innocent questions.
    user_last = req.messages[-1].content if req.messages else ""
    action = data.get("action") or ""

    # start_over → only if user literally typed start over / restart / new forecast
    if action == "start_over" and not re.search(r"start\s*over|restart|new\s*forecast", user_last, re.I):
        action = ""

    # show_parameter_selection → only valid when ALL 6 fields are now filled
    if action == "show_parameter_selection" and still_missing:
        action = ""
    # Auto-trigger show_parameter_selection when all 6 are now filled
    if not still_missing and action not in ("start_over",):
        action = "show_parameter_selection"

    # calculate_forecast / proceed_results / generate_assumptions → only if explicitly requested
    if action == "calculate_forecast" and not re.search(r"calculat|run\s*forecast", user_last, re.I):
        action = ""
    if action == "proceed_results" and not re.search(r"result|chart|show", user_last, re.I):
        action = ""
    if action == "generate_assumptions" and not re.search(r"generat|assumption|next\s*step", user_last, re.I):
        action = ""

    bot_message = data.get("bot_message", "I'm here to help you build your forecast!")

    # ── Guard: if bot_message asks for a field that's no longer missing,
    #    replace the trailing question with a "proceed" prompt. ──────────────
    field_labels = {
        "country": "country", "productName": "product name",
        "classMoa": "class / MoA", "indication": "indication",
        "launchYear": "launch year", "peakYear": "Forecast - End Year",
    }
    if not still_missing:
        # Strip any trailing sentence that asks for a field we already have
        ask_pattern = r"[.!]?\s*(Now[,]?\s*)?(could you|please|can you|what is|what's|provide|tell me)[^.?!]*\??\s*$"
        bot_message = re.sub(ask_pattern, "", bot_message, flags=re.I).strip()
        bot_message += "\n\nAll information collected — proceeding to parameter selection!"
    else:
        # Make sure we aren't asking for a field that IS in field_updates
        for key, label in field_labels.items():
            if key in field_updates and key not in still_missing:
                # Remove any sentence that asks for this now-filled field
                pattern = rf"[^.!?]*\b{re.escape(label)}\b[^.!?]*[.!?]"
                cleaned = re.sub(pattern, "", bot_message, flags=re.I).strip()
                if cleaned:  # only replace if it leaves something meaningful
                    bot_message = cleaned

    return ChatResponse(
        bot_message=bot_message,
        field_updates=field_updates,
        quick_replies=data.get("quick_replies") or [],
        action=action,
    )


# ---------------------------------------------------------------------------
# Secondary-research endpoint  — with real web research
# ---------------------------------------------------------------------------

import urllib.request
import urllib.parse

# ── Credible sources: curated URLs per topic ─────────────────────────────────
# These are fetched at research time to ground LLM outputs in real data.
CREDIBLE_SOURCES: dict[str, list[dict]] = {
    "rheumatoid arthritis": [
        {"title": "ACR – Rheumatoid Arthritis Guideline", "url": "https://www.rheumatology.org/Practice-Quality/Clinical-Support/Clinical-Practice-Guidelines/Rheumatoid-Arthritis", "domain": "rheumatology.org"},
        {"title": "EULAR RA Recommendations", "url": "https://www.eular.org/recommendations_management_ra.cfm", "domain": "eular.org"},
        {"title": "CDC Arthritis Statistics", "url": "https://www.cdc.gov/arthritis/data_statistics/arthritis-related-stats.htm", "domain": "cdc.gov"},
    ],
    "multiple sclerosis": [
        {"title": "MSIF Atlas of MS 2023", "url": "https://www.msif.org/resource/atlas-of-ms/", "domain": "msif.org"},
        {"title": "ECTRIMS/EAN MS Treatment Guidelines", "url": "https://www.ean.org/guidelines", "domain": "ean.org"},
        {"title": "National MS Society Prevalence", "url": "https://www.nationalmssociety.org/About-the-MS-Society/News/New-Prevalence-Data", "domain": "nationalmssociety.org"},
    ],
    "type 2 diabetes": [
        {"title": "IDF Diabetes Atlas 2023", "url": "https://diabetesatlas.org/", "domain": "diabetesatlas.org"},
        {"title": "ADA Standards of Medical Care 2024", "url": "https://diabetesjournals.org/care/issue/47/Supplement_1", "domain": "diabetesjournals.org"},
        {"title": "CDC National Diabetes Statistics", "url": "https://www.cdc.gov/diabetes/data/statistics-report/index.html", "domain": "cdc.gov"},
    ],
    "oncology": [
        {"title": "GLOBOCAN 2022 Cancer Statistics", "url": "https://gco.iarc.fr/today/home", "domain": "gco.iarc.fr"},
        {"title": "NCCN Clinical Practice Guidelines", "url": "https://www.nccn.org/guidelines/category_1", "domain": "nccn.org"},
        {"title": "ASCO Cancer Statistics 2024", "url": "https://ascopubs.org/journal/jco", "domain": "ascopubs.org"},
        {"title": "NCI Cancer Stat Facts", "url": "https://seer.cancer.gov/statfacts/", "domain": "seer.cancer.gov"},
    ],
    "alzheimer disease": [
        {"title": "Alzheimer's Association 2024 Facts & Figures", "url": "https://www.alz.org/alzheimers-dementia/facts-figures", "domain": "alz.org"},
        {"title": "WHO Dementia Statistics", "url": "https://www.who.int/news-room/fact-sheets/detail/dementia", "domain": "who.int"},
        {"title": "AAIC 2023 Prevalence Data", "url": "https://aaic.alz.org/", "domain": "aaic.alz.org"},
    ],
    "heart failure": [
        {"title": "AHA 2024 Heart Disease & Stroke Statistics", "url": "https://www.ahajournals.org/doi/10.1161/CIR.0000000000001123", "domain": "ahajournals.org"},
        {"title": "ESC Heart Failure Guidelines 2021", "url": "https://www.escardio.org/Guidelines/Clinical-Practice-Guidelines/Heart-Failure", "domain": "escardio.org"},
        {"title": "Heart Failure Society of America", "url": "https://hfsa.org/", "domain": "hfsa.org"},
    ],
    # default / fallback
    "_default": [
        {"title": "WHO Global Health Observatory", "url": "https://www.who.int/data/gho", "domain": "who.int"},
        {"title": "NIH National Library of Medicine – PubMed", "url": "https://pubmed.ncbi.nlm.nih.gov/", "domain": "pubmed.ncbi.nlm.nih.gov"},
        {"title": "EMA – European Medicines Agency", "url": "https://www.ema.europa.eu/en", "domain": "ema.europa.eu"},
        {"title": "FDA Drug Approvals & Databases", "url": "https://www.fda.gov/drugs/drug-approvals-and-databases", "domain": "fda.gov"},
    ],
}

# Additional credible domains the LLM is instructed to prefer when citing
CREDIBLE_DOMAINS = [
    "who.int", "cdc.gov", "nih.gov", "pubmed.ncbi.nlm.nih.gov",
    "clinicaltrials.gov", "fda.gov", "ema.europa.eu",
    "evaluate.com",
    "nejm.org", "thelancet.com", "jamanetwork.com", "bmj.com",
    "ahajournals.org", "escardio.org",
    "alz.org", "nationalmssociety.org",
    "diabetesatlas.org", "diabetesjournals.org",
    "rheumatology.org", "eular.org",
    "nccn.org", "ascopubs.org", "gco.iarc.fr", "seer.cancer.gov",
    "msif.org",
]


_research_log = logging.getLogger("research")


def _pubmed_search_with_abstracts(query: str, max_results: int = 3) -> list[dict]:
    """
    Queries PubMed eutils and returns papers WITH full abstract text.
    Uses esearch → esummary (metadata) + efetch (abstracts).
    """
    results = []
    try:
        # Step 1: search for PMIDs
        search_url = (
            "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?"
            + urllib.parse.urlencode({
                "db": "pubmed", "term": query,
                "retmax": max_results, "retmode": "json", "sort": "relevance",
            })
        )
        with urllib.request.urlopen(search_url, timeout=8) as r:
            pmids = json.loads(r.read().decode()).get("esearchresult", {}).get("idlist", [])
        if not pmids:
            return results

        # Step 2: fetch metadata (title, journal, year)
        summary_url = (
            "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esummary.fcgi?"
            + urllib.parse.urlencode({"db": "pubmed", "id": ",".join(pmids), "retmode": "json"})
        )
        with urllib.request.urlopen(summary_url, timeout=8) as r:
            summary = json.loads(r.read().decode()).get("result", {})

        # Step 3: fetch abstracts via efetch XML
        fetch_url = (
            "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi?"
            + urllib.parse.urlencode({"db": "pubmed", "id": ",".join(pmids), "retmode": "xml", "rettype": "abstract"})
        )
        with urllib.request.urlopen(fetch_url, timeout=10) as r:
            xml = r.read().decode("utf-8", errors="ignore")

        # Extract abstract texts from XML (ordered, matches pmid order)
        raw_abstracts = re.findall(r'<AbstractText[^>]*>(.*?)</AbstractText>', xml, re.DOTALL)
        abstract_map: dict[str, str] = {}
        for i, pmid in enumerate(pmids):
            parts = raw_abstracts[i * 1:(i + 1) * 4]  # up to 4 labeled sections per article
            combined = " ".join(re.sub(r'<[^>]+>', '', p) for p in parts).strip()
            if combined:
                abstract_map[pmid] = combined[:800]  # cap at 800 chars

        for pmid in pmids:
            art = summary.get(pmid, {})
            title = art.get("title", "").strip()
            if not title:
                continue
            results.append({
                "title": title,
                "url": f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/",
                "domain": "pubmed.ncbi.nlm.nih.gov",
                "source": art.get("source", ""),
                "year": art.get("pubdate", "")[:4],
                "pmid": pmid,
                "abstract": abstract_map.get(pmid, ""),   # ← REAL abstract text
            })
    except Exception as exc:
        _research_log.warning(f"PubMed search failed for '{query}': {exc}")
    return results


def _fetch_page_text(url: str, max_chars: int = 900) -> str:
    """
    Fetch a webpage and extract the most statistically dense plain-text snippet.
    Strips all HTML/script/style tags, then scores 200-word windows by presence
    of numbers, percentages, and epidemiology keywords.
    Returns '' on any error (timeouts, JS-only pages, auth walls).
    """
    try:
        req = urllib.request.Request(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 (compatible; ForecastResearchBot/1.0)",
                "Accept": "text/html,application/xhtml+xml",
                "Accept-Language": "en-US,en;q=0.9",
            },
        )
        with urllib.request.urlopen(req, timeout=7) as r:
            raw = r.read(120_000).decode("utf-8", errors="ignore")  # read max 120 KB

        # Strip scripts, styles, nav elements
        raw = re.sub(r'(?s)<(script|style|nav|footer|header)[^>]*>.*?</\1>', ' ', raw, flags=re.I)
        text = re.sub(r'<[^>]+>', ' ', raw)          # strip remaining tags
        text = re.sub(r'&[a-z#0-9]+;', ' ', text)   # decode basic HTML entities
        text = re.sub(r'\s+', ' ', text).strip()

        # Score 40-word windows; pick the densest data passage
        STAT_KWS = re.compile(
            r'\b(?:prevalence|incidence|patients|diagnosis|diagnosed|treatment|rate|'
            r'percent|share|cost|price|annual|billion|million|thousand|mortality|'
            r'survival|efficacy|eligible|biomarker)\b',
            re.I,
        )
        NUM_PAT = re.compile(r'\d[\d,]*\.?\d*\s*(?:%|percent|million|billion|thousand)?', re.I)
        words = text.split()
        window = 60  # words per scored window
        best, best_score = "", 0
        for i in range(0, min(len(words), 4000), 20):
            chunk = " ".join(words[i: i + window])
            score = len(STAT_KWS.findall(chunk)) * 2 + len(NUM_PAT.findall(chunk))
            if score > best_score:
                best_score, best = score, " ".join(words[i: i + window * 2])
        return best[:max_chars] if best else text[:max_chars]
    except Exception as exc:
        _research_log.warning(f"Page fetch failed for {url}: {exc}")
        return ""


def _get_curated_sources(indication: str) -> list[dict]:
    """Returns curated source list for a given indication (case-insensitive)."""
    ind_lower = (indication or "").lower()
    for key, sources in CREDIBLE_SOURCES.items():
        if key in ind_lower or ind_lower in key:
            return sources
    return CREDIBLE_SOURCES["_default"]


async def _fetch_web_research_context(indication: str, country: str, class_moa: str) -> tuple[str, list[dict]]:
    """
    1. Runs 5 focused PubMed searches → fetches real abstracts via efetch.
    2. Attempts to scrape content from 3 curated authoritative pages.
    3. Builds a rich context block fed directly into the LLM prompt.
    Returns (context_text, all_sources).
    """
    # ── PubMed searches (parallel via asyncio.gather) ──────────────────────
    queries = [
        f"{indication} prevalence {country}",
        f"{indication} incidence epidemiology statistics",
        f"{indication} diagnosis treatment rate {country}",
        f"{class_moa} {indication} market share adoption",
        f"{indication} annual cost therapy pricing {country}",
    ]
    search_tasks = [
        asyncio.to_thread(_pubmed_search_with_abstracts, q, 2)
        for q in queries
    ]
    all_pubmed_batches = await asyncio.gather(*search_tasks, return_exceptions=True)

    seen_pmids: set[str] = set()
    unique_pubmed: list[dict] = []
    for batch in all_pubmed_batches:
        if isinstance(batch, Exception):
            continue
        for art in batch:
            pmid = art.get("pmid", "")
            if pmid and pmid not in seen_pmids:
                seen_pmids.add(pmid)
                unique_pubmed.append(art)

    # ── Scrape curated authoritative pages (up to 3, parallel) ─────────────
    curated = _get_curated_sources(indication)
    page_tasks = [
        asyncio.to_thread(_fetch_page_text, src["url"], 900)
        for src in curated[:3]
    ]
    page_contents = await asyncio.gather(*page_tasks, return_exceptions=True)

    # ── Build LLM context string ────────────────────────────────────────────
    lines = [f"=== Real Research Content for: {indication} | {country} | {class_moa} ===\n"]

    # PubMed abstracts section
    pubmed_with_abstracts = [a for a in unique_pubmed if a.get("abstract")]
    if pubmed_with_abstracts:
        lines.append("── PubMed Articles (with abstract text) ──")
        for art in pubmed_with_abstracts[:6]:
            lines.append(
                f"\n[{art['source']} {art['year']}] {art['title']}\n"
                f"URL: {art['url']}\n"
                f"Abstract: {art['abstract']}"
            )
    elif unique_pubmed:
        lines.append("── PubMed Articles (titles only — abstracts unavailable) ──")
        for art in unique_pubmed[:6]:
            lines.append(f"  • {art['title']} ({art['source']} {art['year']}) — {art['url']}")

    # Curated page content section
    lines.append("\n── Authoritative Source Pages (fetched content) ──")
    for src, content in zip(curated[:3], page_contents):
        if isinstance(content, Exception) or not content:
            lines.append(f"  • {src['title']} — {src['url']} [content unavailable]")
        else:
            lines.append(
                f"\n[{src['domain']}] {src['title']}\n"
                f"URL: {src['url']}\n"
                f"Extracted data: {content}"
            )

    lines.append(
        "\n── Additional curated sources (not scraped) ──\n"
        + "\n".join(f"  • {s['title']} — {s['url']}" for s in curated[3:])
    )

    all_sources = curated + unique_pubmed
    return "\n".join(lines), all_sources


RESEARCH_SYSTEM = """\
You are a senior commercial pharmaceutical analyst with deep expertise in epidemiology,
market access, and drug pricing across major markets.

You will receive REAL content fetched from PubMed abstracts and authoritative web pages
(WHO, CDC, AHA, IDF, GLOBOCAN, etc.). Prefer numbers from that fetched content when
available. When the fetched content is unavailable, unhelpful, or lacks relevant data
for a specific field, draw on your expert training knowledge to provide a well-reasoned
estimate — you are a senior pharma analyst and your knowledge is a valid source.

Respond ONLY with a valid JSON object matching this exact schema (no extra keys, no prose):
{
  "prevalence":           <float 0–1, proportion of population — extract from abstract/page data>,
  "prevalenceRationale":  "<cite fetched source + URL if available; otherwise provide a well-reasoned explanation>",
  "diagnosis":            <float 0–1>,
  "diagnosisRationale":   "<cite fetched source + URL if available; otherwise provide a well-reasoned explanation>",
  "treatment":            <float 0–1>,
  "treatmentRationale":   "<cite fetched source + URL if available; otherwise provide a well-reasoned explanation>",
  "biomarker":            <float 0–1, eligibility / biomarker-positive rate>,
  "biomarkerRationale":   "<cite fetched source + URL if available; otherwise provide a well-reasoned explanation>",
  "classShare":           <float 0–1, drug-class peak market share>,
  "classShareRationale":  "<cite fetched source + URL if available; otherwise provide a well-reasoned explanation>",
  "productShare":         <float 0–1, product share within class at peak>,
  "productShareRationale":"<cite fetched source + URL if available; otherwise provide a well-reasoned explanation>",
  "annualCost":           <integer USD, annual cost per patient>,
  "costRationale":        "<cite fetched source + URL if available; otherwise provide a well-reasoned explanation>",
  "discountRationale":    "<cite fetched source + URL if available; otherwise provide a well-reasoned explanation>",
  "sources": [
    { "title": "<paper / guideline title>", "url": "<full URL>", "domain": "<hostname>" },
    ...
  ]
}

Rules:
• All rates/shares are proportions (0–1), NOT percentages.
• DERIVE numeric values from the abstract and page text provided — if a specific number
  appears in the fetched content (e.g. "prevalence of 0.55%"), use it directly.
• In every rationale field: quote or closely paraphrase the relevant sentence from the
  fetched content, then append " — <full URL from the context>". Never invent URLs.
• If the fetched content does not contain a relevant number for a field, use your
  expert knowledge to construct a well-reasoned rationale. Explain the reasoning
  (e.g., typical ranges for this indication, comparable market benchmarks, published
  guidelines). Do NOT just say "no data retrieved" — always provide a substantive
  rationale with a numeric estimate.
• Include 4–8 sources in the sources array drawn from the research context URLs.
• Adjust annualCost and discountRationale to the specified country's pricing environment.
"""


@app.post("/api/research")
async def research(req: ResearchRequest):
    # Step 1: Fetch real web research context
    context_text, all_sources = await _fetch_web_research_context(
        req.indication, req.country, req.class_moa
    )

    prompt = (
        f"Indication:    {req.indication}\n"
        f"Country:       {req.country}\n"
        f"Drug Class/MoA:{req.class_moa}\n\n"
        f"{context_text}\n\n"
        "Using the research context above, generate commercial forecast assumptions for this asset. "
        "Cite the actual sources provided in the rationale fields."
    )

    resp = await asyncio.to_thread(
        _bedrock.invoke_with_retry,
        messages=[{"role": "user", "content": prompt}],
        system_prompt=RESEARCH_SYSTEM,
        temperature=0.3,
        max_tokens=2000,
    )

    raw = resp["content"]
    # Strip markdown code fences if the LLM wrapped the JSON
    cleaned = re.sub(r"^```(?:json)?\s*", "", raw.strip(), flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*```$", "", cleaned.strip())
    m = re.search(r"\{[\s\S]*\}", cleaned)
    if m:
        cleaned = m.group(0)

    try:
        result = json.loads(cleaned)
        # Always ensure curated sources are included
        existing_urls = {s.get("url") for s in result.get("sources", [])}
        curated = _get_curated_sources(req.indication)
        for src in curated[:3]:
            if src["url"] not in existing_urls:
                result.setdefault("sources", []).append(src)
                existing_urls.add(src["url"])
        # Add top PubMed hits not already in list
        for src in all_sources:
            if src.get("pmid") and src["url"] not in existing_urls:
                result.setdefault("sources", []).append({
                    "title": src["title"],
                    "url": src["url"],
                    "domain": src["domain"],
                })
                existing_urls.add(src["url"])
                if len(result["sources"]) >= 8:
                    break
        return result
    except json.JSONDecodeError:
        return {"error": "Failed to parse research data", "raw": raw}


# ---------------------------------------------------------------------------
# Parameter recommendation endpoint
# ---------------------------------------------------------------------------

RECOMMEND_SYSTEM = """\
You are a commercial forecasting expert specialising in pharmaceutical revenue models.

Given a product's details, recommend which forecast flow parameters to include and explain why.

Available parameter IDs and their meaning:
  prevalence          – use for chronic / prevalent diseases (RA, MS, T2D, HF, AD)
  incidence           – use for oncology or acute diseases (use INSTEAD of prevalence)
  diagnosisRate       – proportion of patients who are diagnosed
  severity            – disease severity subtype filter
  treatmentRate       – proportion of diagnosed patients who receive treatment
  eligibilityCriteria – biomarker positivity, line-of-therapy, or inclusion criteria
  progressionRate     – disease progression or line-advancement rate
  classShare          – drug-class peak market share
  peakProductShare    – product share within the class at peak
  annualCostPerPatient – annual gross treatment cost per patient
  discount            – rebate / net pricing discount rate

Rules:
  - Include EITHER prevalence OR incidence, never both.
  - Always include classShare, peakProductShare, annualCostPerPatient, discount.
  - Always include diagnosisRate unless the indication has near-universal diagnosis.
  - For oncology: use incidence, include eligibilityCriteria (biomarker), exclude severity.
  - For rare disease: use prevalence, include eligibilityCriteria; treatmentRate optional.
  - For chronic disease (RA, MS, T2D, HF): use prevalence, include treatmentRate.

Respond ONLY with a valid JSON object (no markdown fences, no prose outside it):
{
  "recommendation": "<4-5 sentence narrative. Sentence 1: therapy area profile and WHY you chose prevalence vs incidence for this indication. Sentence 2: which patient attrition parameters were selected and why they matter most for this specific asset. Sentence 3: any eligibility, biomarker, or severity filters and their rationale. Sentence 4: market sizing parameters and what drives the biggest uncertainty. Sentence 5: the single parameter that most determines peak revenue for this asset. Use **bold** for every parameter name.>",
  "epi_note": "<1 sentence: concise reason for choosing prevalence OR incidence for this indication>",
  "flow_note": "<1 sentence: which patient-flow attrition steps dominate and why>",
  "pricing_note": "<1 sentence: pricing and competitive share dynamics specific to this therapy area>",
  "params": ["<paramId1>", "<paramId2>", ...]
}
"""


class RecommendRequest(BaseModel):
    indication: str
    product_name: str = ""
    class_moa: str = ""
    country: str = ""


@app.post("/api/recommend")
async def recommend(req: RecommendRequest):
    prompt = (
        f"Product:    {req.product_name or 'unnamed'}\n"
        f"Class/MoA:  {req.class_moa or 'unknown'}\n"
        f"Indication: {req.indication or 'unspecified'}\n"
        f"Country:    {req.country or 'unspecified'}\n\n"
        "Recommend the optimal set of forecast flow parameters for this asset."
    )

    resp = await asyncio.to_thread(
        _bedrock.invoke_with_retry,
        messages=[{"role": "user", "content": prompt}],
        system_prompt=RECOMMEND_SYSTEM,
        temperature=0.3,
        max_tokens=700,
    )

    raw = resp["content"]
    # Strip markdown code fences if the LLM wrapped the JSON
    cleaned = re.sub(r"^```(?:json)?\s*", "", raw.strip(), flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*```$", "", cleaned.strip())
    # Also try to extract the first {...} block in case of surrounding prose
    m = re.search(r"\{[\s\S]*\}", cleaned)
    if m:
        cleaned = m.group(0)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        return {"error": "Failed to parse recommendation", "raw": raw}


# ---------------------------------------------------------------------------
# Forecast computation — single consolidated function
# Patient flow: Population → Affected (prevalence/incidence) → Diagnosed → Eligible
#   → [Severity] → Treated → [Progression] → Class → Product → Revenue
# ---------------------------------------------------------------------------

def compute_forecast(assumptions: dict, selected_parameters: list) -> dict:
    """
    Single consolidated forecast calculation. Correctly applies all selected indicators.

    Patient funnel (each step uses corresponding param when selected):
      population → prevalence OR incidence → diagnosisRate → eligibilityCriteria
      → severity → treatmentRate → progressionRate → classShare → peakProductShare
      → revenue (annualCostPerPatient, discount)

    Handles: unit conversion (rate, per100k, per1M), YoY growth, linear/S-curve adoption.
    """

    print("assumptions", assumptions)
    print("selected_parameters", selected_parameters)

    params = set(selected_parameters or [])

    def get_val(obj) -> float:
        if obj is None:
            return 0.0
        if isinstance(obj, (int, float)):
            return float(obj)
        if not isinstance(obj, dict):
            return 0.0
        return float(obj.get("value", 0) or 0)

    def get_val_for_year(obj, i: int) -> float:
        v = get_val(obj)
        if isinstance(obj, dict) and obj.get("yoyGrowth") is not None:
            try:
                r = float(obj["yoyGrowth"])
                return v * ((1 + r) ** i)
            except (TypeError, ValueError):
                pass
        return v

    def to_rate(obj) -> float:
        """Convert prevalence/incidence to proportion (0-1). Handles rate, per100k, per1M."""
        v = get_val(obj)
        if not isinstance(obj, dict):
            return v
        ut = obj.get("unitType") or "rate"
        if ut == "per100k":
            return v / 100_000.0
        if ut == "per1M":
            return v / 1_000_000.0
        return v

    def to_rate_for_year(obj, i: int) -> float:
        if obj is None or not isinstance(obj, dict):
            return 0.0
        v = get_val(obj)
        ut = obj.get("unitType") or "rate"
        divisor = 100_000.0 if ut == "per100k" else (1_000_000.0 if ut == "per1M" else 1.0)
        if obj.get("yoyGrowth") is not None:
            try:
                r = float(obj["yoyGrowth"])
                v = v * ((1 + r) ** i)
            except (TypeError, ValueError):
                pass
        return v / divisor

    def ramp_linear(i: int, ttp: int) -> float:
        if ttp <= 0:
            return 1.0
        return min(1.0, (i + 1) / ttp)

    def ramp_scurve(i: int, ttp: int, start: float, peak: float) -> float:
        if i <= 0:
            return start
        ttp = max(1, ttp)
        if i >= ttp:
            return peak
        tr = i / ttp
        # Normalize sigmoid so tr=0→start and tr=1→peak exactly.
        _s0 = 1.0 / (1.0 + math.exp(2.5))   # sigmoid at tr=0
        _s1 = 1.0 / (1.0 + math.exp(-2.5))  # sigmoid at tr=1
        sig = 1.0 / (1.0 + math.exp(-5.0 * (tr - 0.5)))
        t = (sig - _s0) / (_s1 - _s0)
        return start + (peak - start) * t

    launch_year = int(assumptions.get("launchYear") or assumptions.get("launch_year") or 2026)
    peak_year = int(assumptions.get("peakYear") or assumptions.get("peak_year") or launch_year + 5)

    pop_obj = assumptions.get("population") or {}
    pop_growth = float(pop_obj.get("yoyGrowth", 0) or 0) if isinstance(pop_obj, dict) else 0.0
    base_pop = get_val(pop_obj)

    prev_obj = assumptions.get("prevalence") if "prevalence" in params else (assumptions.get("incidence") if "incidence" in params else None)
    diag_obj = assumptions.get("diagnosisRate") or {}
    elig_obj = assumptions.get("eligibilityCriteria") or {}
    sev_obj = assumptions.get("severity") or {}
    treat_obj = assumptions.get("treatmentRate") or {}
    prog_obj = assumptions.get("progressionRate") or {}
    cs_obj = assumptions.get("classShare") or assumptions.get("class_share") or {}
    ps_obj = assumptions.get("peakProductShare") or assumptions.get("peak_product_share") or {}
    ac_obj = assumptions.get("annualCostPerPatient") or assumptions.get("annual_cost_per_patient") or assumptions.get("annualCost") or {}
    dr_obj = assumptions.get("discount") or {}

    def _ttp(obj: dict) -> int:
        if isinstance(obj, dict):
            if obj.get("timeToPeak"):
                return max(1, int(obj["timeToPeak"]))
            if obj.get("peakYear"):
                return max(1, int(obj["peakYear"]) - launch_year)
        return max(1, peak_year - launch_year)

    cs_ttp = _ttp(cs_obj)
    ps_ttp = _ttp(ps_obj)
    cs_ss = float(cs_obj.get("startingShare", 0.05) or 0.05) if isinstance(cs_obj, dict) else 0.05
    ps_ss = float(ps_obj.get("startingShare", 0.03) or 0.03) if isinstance(ps_obj, dict) else 0.03
    cs_ct = (cs_obj.get("curveType") or "scurve").lower() if isinstance(cs_obj, dict) else "scurve"
    ps_ct = (ps_obj.get("curveType") or "scurve").lower() if isinstance(ps_obj, dict) else "scurve"
    cs_peak = get_val(cs_obj)
    ps_peak = get_val(ps_obj)

    results = []
    max_ns, max_nsy, max_p, max_gs = 0.0, 0, 0, 0.0
    prev_pp, prev_net = 0.0, 0.0

    for i in range(peak_year - launch_year + 1):
        year = launch_year + i

        total_pop = base_pop * ((1 + pop_growth) ** i) if base_pop else 0.0

        if prev_obj and ("prevalence" in params or "incidence" in params):
            epi_rate = to_rate_for_year(prev_obj, i) if (isinstance(prev_obj, dict) and prev_obj.get("yoyGrowth") is not None) else to_rate(prev_obj)
            affected = total_pop * epi_rate
        else:
            affected = total_pop

        diag_val = get_val_for_year(diag_obj, i) if "diagnosisRate" in params else 1.0
        diagnosed = affected * diag_val

        elig_val = get_val_for_year(elig_obj, i) if "eligibilityCriteria" in params else 1.0
        eligible = diagnosed * elig_val

        sev_val = get_val_for_year(sev_obj, i) if "severity" in params and sev_obj else 1.0
        eligible_for_treat = eligible * sev_val

        treat_val = get_val_for_year(treat_obj, i) if "treatmentRate" in params else 1.0
        treated = eligible_for_treat * treat_val

        prog_val = get_val_for_year(prog_obj, i) if "progressionRate" in params and prog_obj else 0.0
        treated *= max(0.0, 1.0 - prog_val)

        if cs_ct == "linear":
            cs = cs_peak * ramp_linear(i, cs_ttp)
        else:
            cs = ramp_scurve(i, cs_ttp, cs_ss, cs_peak)
        class_patients = treated * cs

        if ps_ct == "linear":
            ps = ps_peak * ramp_linear(i, ps_ttp)
        else:
            ps = ramp_scurve(i, ps_ttp, ps_ss, ps_peak)
        product_patients = class_patients * ps
        product_patients_int = round(product_patients)

        ac_val = get_val_for_year(ac_obj, i)
        dr_val = get_val(dr_obj)
        gross_revenue = product_patients * ac_val
        gross_sales_m = gross_revenue / 1e6
        net_sales_m = gross_sales_m * (1 - dr_val)

        patient_growth_pct = ((product_patients - prev_pp) / prev_pp) if prev_pp > 0 else 0.0
        revenue_growth_pct = ((net_sales_m - prev_net) / prev_net) if prev_net > 0 else 0.0
        prev_pp, prev_net = product_patients, net_sales_m

        results.append({
            "year": year,
            "totalPopulation": round(total_pop),
            "affectedPatients": round(affected),
            "diagnosedPatients": round(diagnosed),
            "eligiblePatients": round(eligible),
            "treatedMarket": round(treated),
            "classPatients": round(class_patients),
            "treatedPatients": product_patients_int,
            "productPatients": product_patients_int,
            "classShare": f"{(cs * 100):.1f}",
            "productShare": f"{(ps * 100):.1f}",
            "annualCost": f"{int(ac_val):,}",
            "grossSales": f"{gross_sales_m:.1f}",
            "discount": f"{(dr_val * 100):.0f}",
            "netSales": f"{net_sales_m:.1f}",
            "patientGrowthPct": patient_growth_pct,
            "revenueGrowthPct": revenue_growth_pct,
        })
        if net_sales_m > max_ns:
            max_ns, max_nsy, max_p, max_gs = net_sales_m, year, product_patients_int, gross_sales_m

    return {
        "forecast_results": results,
        "peak_net_sales": max_ns,
        "peak_year": max_nsy,
        "peak_patients": max_p,
        "peak_gross_sales": max_gs,
    }


class ForecastRequest(BaseModel):
    assumptions: dict = {}
    selected_parameters: list = []


@app.post("/api/forecast")
async def forecast_compute(req: ForecastRequest):
    """Compute forecast table and dashboard summary from base assumptions."""
    try:
        return compute_forecast(req.assumptions, req.selected_parameters)
    except Exception as exc:
        logging.getLogger("ForecastAPI").error("Forecast compute error: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail=str(exc))


@app.post("/api/sensitivity")
async def sensitivity_analysis(req: ForecastRequest):
    """Tornado chart: vary each key assumption ±20%, return impact on peak net sales."""
    import copy
    try:
        base = compute_forecast(req.assumptions, req.selected_parameters)
        base_peak = base["peak_net_sales"]

        # Map param key → display label
        PARAM_LABELS = {
            "prevalence":           "Prevalence / Incidence",
            "incidence":            "Prevalence / Incidence",
            "diagnosisRate":        "Diagnosis Rate",
            "eligibilityCriteria":  "Eligibility Criteria",
            "treatmentRate":        "Treatment Rate",
            "classShare":           "Class Share",
            "peakProductShare":     "Product Share",
            "annualCostPerPatient": "Annual Cost / Patient",
            "annualCost":           "Annual Cost / Patient",
            "discount":             "Discount Rate",
        }

        seen_labels: set = set()
        results = []

        for param, label in PARAM_LABELS.items():
            if param not in req.assumptions:
                continue
            if label in seen_labels:
                continue

            obj = req.assumptions[param]
            if not isinstance(obj, dict) or obj.get("value") is None:
                continue

            base_val = float(obj.get("value") or 0)
            if base_val == 0:
                continue

            row = {"param": param, "label": label}
            for delta, key in [(-0.2, "low"), (0.2, "high")]:
                tweaked = copy.deepcopy(req.assumptions)
                tweaked[param] = {**tweaked[param], "value": base_val * (1 + delta)}
                scenario = compute_forecast(tweaked, req.selected_parameters)
                row[key] = round(scenario["peak_net_sales"], 2)

            row["impact"] = round(row["high"] - row["low"], 2)
            results.append(row)
            seen_labels.add(label)

        results.sort(key=lambda r: r["impact"], reverse=True)
        return {"base_peak": round(base_peak, 2), "sensitivity": results}

    except Exception as exc:
        logging.getLogger("ForecastAPI").error("Sensitivity error: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail=str(exc))


# ---------------------------------------------------------------------------
# Config (frontend uses this to skip save-input when S3 enabled)
# ---------------------------------------------------------------------------
@app.get("/api/config")
async def get_config():
    """Return config flags for frontend (e.g. skip form persistence when S3 used)."""
    return {"s3_enabled": s3_enabled()}


# ---------------------------------------------------------------------------
# Save user input (local only when S3 disabled; never writes to S3)
# ---------------------------------------------------------------------------
@app.post("/api/save-input")
async def save_input(data: dict):
    """Persist user inputs locally only when S3 disabled. When S3 enabled, do nothing."""
    if s3_enabled():
        return {"status": "saved", "path": "s3"}
    try:
        from config import PROJECT_ROOT
        output_path = PROJECT_ROOT / "data" / "user_input.json"
    except ImportError:
        output_path = _project_root / "data" / "user_input.json"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    return {"status": "saved", "path": str(output_path)}


# ---------------------------------------------------------------------------
# Agent endpoint — delegates to the LangGraph orchestrator in ../../../agent/
# ---------------------------------------------------------------------------

# Lazy-import the orchestrator so the server still starts even if the agent
# directory is unavailable (graceful degradation).
_orchestrator_cls = None
_agent_sessions: Dict[str, object] = {}
_agent_workbook_info: Dict[str, dict] = {}  # session_id -> {local, s3_key}
_agent_run_dirs: Dict[str, Path] = {}  # session_id -> run_dir (logs + outputs + scripts)
_agent_job_status: Dict[str, dict] = {}  # session_id -> {status, response?, error?, workbook_path?}

# PPTX agent state (parallel to Excel agent)
_pptx_job_status: Dict[str, dict] = {}   # session_id -> {status, pptx_path?, error?}
_pptx_file_info: Dict[str, dict] = {}    # session_id -> {local, s3_key}

def _generate_run_session_id() -> str:
    """Session ID with timestamp first for chronological sorting."""
    return f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"


def _get_agent_orchestrator(session_id: str, run_dir: Path):
    global _orchestrator_cls
    if _orchestrator_cls is None:
        try:
            from agents.excel_agent import ExcelAgentOrchestrator
            _orchestrator_cls = ExcelAgentOrchestrator
        except Exception as exc:
            raise RuntimeError(f"Could not import agent orchestrator: {exc}") from exc

    if session_id not in _agent_sessions:
        try:
            from config import PROMPTS_DIR
            prompts_path = PROMPTS_DIR
        except ImportError:
            prompts_path = _agent_dir / "prompts"
        run_dir.mkdir(parents=True, exist_ok=True)
        _agent_sessions[session_id] = _orchestrator_cls(
            aws_region=os.getenv("AWS_REGION", "us-east-1"),
            prompts_dir=str(prompts_path),
            run_dir=run_dir,
            session_id=session_id,
        )
    return _agent_sessions[session_id]


class AgentRequest(BaseModel):
    session_id: str = ""
    user_input: dict = {}


def _run_excel_agent_task(session_id: str, run_dir: Path, prompt: str, user_input: dict):
    """Runs in thread pool; updates _agent_job_status and _agent_workbook_info on completion."""
    _use_s3 = s3_enabled()
    try:
        run_dir.mkdir(parents=True, exist_ok=True)
        if _use_s3:
            upload_json(user_input, s3_logs_key(session_id, "user_input.json"))
        else:
            user_input_path = run_dir / "user_input.json"
            with open(user_input_path, "w", encoding="utf-8") as f:
                json.dump(user_input, f, indent=2, ensure_ascii=False)
        orchestrator = _get_agent_orchestrator(session_id, run_dir)
        orig = os.getcwd()
        try:
            os.chdir(str(_agent_dir))
            response = orchestrator.process_request(prompt)
        finally:
            os.chdir(orig)
        workbook_path = None
        if response and "Path:" in response:
            for line in response.split("\n"):
                line = line.strip()
                if line.startswith("Path:"):
                    workbook_path = line[5:].strip()
                    break
        info = {"local": workbook_path, "s3_key": None}
        if workbook_path and _use_s3 and upload_file:
            key = s3_output_key(session_id, Path(workbook_path).name)
            if upload_file(Path(workbook_path), key):
                info["s3_key"] = key
        if workbook_path:
            _agent_workbook_info[session_id] = info
        job = {
            "status": "done",
            "response": response,
            "workbook_path": workbook_path,
            "success": True,
        }
        _agent_job_status[session_id] = job
        try:
            with open(run_dir / "job_status.json", "w", encoding="utf-8") as f:
                json.dump(job, f)
        except Exception:
            pass
    except Exception as exc:
        logging.getLogger("AgentEndpoint").error("Agent error: %s", exc, exc_info=True)
        job = {"status": "error", "error": str(exc), "success": False}
        _agent_job_status[session_id] = job
        try:
            with open(run_dir / "job_status.json", "w", encoding="utf-8") as f:
                json.dump(job, f)
        except Exception:
            pass
    finally:
        # Ensure buffered logs are uploaded before the temp dir is deleted.
        try:
            orch = _agent_sessions.get(session_id)
            if orch is not None:
                orch._session_manager.flush_logs()
        except Exception:
            pass
        if _use_s3 and run_dir.exists():
            shutil.rmtree(run_dir, ignore_errors=True)

def _dict_to_compact_str(obj, indent: int = 0) -> str:
    """Convert dict/list to compact string. Preserves all keys, minimal syntax (no braces/quotes)."""
    pad = "  " * indent

    def scalar(v):
        if v is None:
            return "null"
        if isinstance(v, bool):
            return "true" if v else "false"
        return str(v)

    if isinstance(obj, dict):
        lines = []
        for k, v in obj.items():
            if isinstance(v, dict) and v:
                lines.append(f"{pad}{k}:")
                lines.append(_dict_to_compact_str(v, indent + 1))
            elif isinstance(v, list) and v:
                lines.append(f"{pad}{k}:")
                lines.append(_dict_to_compact_str(v, indent + 1))
            else:
                val = _dict_to_compact_str(v, indent + 1) if isinstance(v, (dict, list)) else scalar(v)
                lines.append(f"{pad}{k}: {val}")
        return "\n".join(lines)
    if isinstance(obj, list):
        lines = []
        for i, item in enumerate(obj):
            if isinstance(item, (dict, list)) and (item if not isinstance(item, list) else len(item)):
                lines.append(f"{pad}{i}:")
                lines.append(_dict_to_compact_str(item, indent + 1))
            else:
                val = _dict_to_compact_str(item, indent + 1) if isinstance(item, (dict, list)) else scalar(item)
                lines.append(f"{pad}{i}: {val}")
        return "\n".join(lines)
    return scalar(obj)


@app.post("/api/agent")
async def run_agent(req: AgentRequest):
    """Start the LangGraph Excel agent in background; returns immediately. Poll GET /api/agent/status."""
    session_id = _generate_run_session_id()
    if s3_enabled():
        run_dir = Path(tempfile.mkdtemp(prefix=f"fa_{session_id}_"))
    else:
        try:
            from config import RUNS_DIR
        except ImportError:
            RUNS_DIR = _project_root / "data" / "runs"
        run_dir = RUNS_DIR / session_id

    forecast_results = req.user_input.get("forecast_results", [])
    ui = {k: v for k, v in req.user_input.items() if k != "forecast_results"}
    user_input_text = _dict_to_compact_str(ui)
    if forecast_results:
        results_text = _dict_to_compact_str(forecast_results)
        prompt = (
            f"Create a forecast for this data\n\n{user_input_text}\n\n"
            f"Pre-calculated forecast results — use these exact numbers in the Excel workbook "
            f"(do not recalculate; these are the authoritative values):\n{results_text}"
        )
    else:
        prompt = f"Create a forecast for this data\n\n{user_input_text}"

    job = _agent_job_status.get(session_id)
    if job and job.get("status") == "running":
        return {"session_id": session_id, "status": "started", "message": "Agent already running"}
    _agent_job_status[session_id] = {"status": "running", "started_at": time.time()}
    _agent_run_dirs[session_id] = run_dir
    asyncio.create_task(asyncio.to_thread(_run_excel_agent_task, session_id, run_dir, prompt, ui))
    return {"session_id": session_id, "status": "started", "success": True}


@app.get("/api/agent/status")
async def agent_status(session_id: str):
    """Poll for Excel agent completion. Returns status: running|done|error."""
    job = _agent_job_status.get(session_id)
    if not job:
        # Server may have restarted — try reading persisted status from disk
        try:
            from config import RUNS_DIR
        except ImportError:
            RUNS_DIR = _project_root / "data" / "runs"
        status_file = RUNS_DIR / session_id / "job_status.json"
        if status_file.is_file():
            try:
                with open(status_file, encoding="utf-8") as f:
                    job = json.load(f)
                _agent_job_status[session_id] = job  # restore to memory
            except Exception:
                pass
    if not job:
        return {"session_id": session_id, "status": "unknown", "success": False}
    st = job.get("status", "unknown")
    out = {"session_id": session_id, "status": st}
    if st == "running":
        started = job.get("started_at", time.time())
        out["started_at"] = started
        out["elapsed_s"] = time.time() - started
    elif st == "done":
        out["success"] = job.get("success", True)
        out["response"] = job.get("response")
        out["workbook_path"] = job.get("workbook_path")
        out["started_at"] = job.get("started_at")
    elif st == "error":
        out["success"] = False
        out["error"] = job.get("error", "Unknown error")
    return out

# ── Excel data parsing helpers (for /api/excel/data preview endpoint) ────────

_EXCEL_EPOCH = datetime(1899, 12, 30)
_DATE_TOKEN_RE = re.compile(
    r'(?<![#0,])(?:y{1,4}|d{1,4}(?![eE])|h{1,2}(?!\])|AM/PM|am/pm)',
    re.IGNORECASE,
)

def _is_date_format(fmt: str) -> bool:
    if not fmt or fmt in ("General", "@", ""):
        return False
    clean = re.sub(r'"[^"]*"|\[[^\]]*\]', "", fmt)
    return bool(_DATE_TOKEN_RE.search(clean))

def _datetime_to_serial(dt) -> float:
    if isinstance(dt, datetime):
        delta = dt - _EXCEL_EPOCH
        return delta.days + delta.seconds / 86400.0 + delta.microseconds / 86400e6
    return (date(dt.year, dt.month, dt.day) - date(1899, 12, 30)).days

def _format_number_excel(value, fmt):
    if value is None:
        return ""
    try:
        if "%" in fmt:
            decimals = len(re.search(r'\.?(0*)', fmt.split("%")[0]).group(1))
            return f"{value * 100:.{decimals}f}%"
        if "$" in fmt or "£" in fmt or "€" in fmt:
            symbol = re.search(r'[\$£€]', fmt).group()
            decimals = 2
            m = re.search(r'\.0+', fmt)
            if m:
                decimals = len(m.group()) - 1
            return f"{symbol}{value:,.{decimals}f}"
        if "," in fmt:
            m = re.search(r'\.?(0*)', fmt.split(",")[-1])
            decimals = len(m.group(1)) if m else 0
            return f"{value:,.{decimals}f}"
        m = re.search(r'\.(0+)', fmt)
        if m:
            return f"{value:.{len(m.group(1))}f}"
    except Exception:
        pass
    if isinstance(value, float) and value == int(value) and abs(value) < 1e15:
        return str(int(value))
    return str(value)

def _cell_display_excel(raw, number_format):
    if raw is None:
        return ""
    if isinstance(raw, bool):
        return "TRUE" if raw else "FALSE"
    if isinstance(raw, datetime):
        return raw.strftime("%m/%d/%Y %H:%M") if (raw.second or raw.minute or raw.hour) else raw.strftime("%m/%d/%Y")
    if isinstance(raw, date):
        return raw.strftime("%m/%d/%Y")
    if isinstance(raw, (int, float)):
        fmt = number_format or "General"
        if fmt in ("General", "@", ""):
            if isinstance(raw, float) and raw == int(raw) and abs(raw) < 1e15:
                return str(int(raw))
            return str(raw)
        return _format_number_excel(raw, fmt)
    return str(raw)

def _rgb_excel(color_obj):
    try:
        if color_obj.type == "rgb":
            v = color_obj.rgb
            if v and v.upper() not in ("00000000", "FF000000", "FFFFFFFF"):
                return "#" + v[-6:]
    except Exception:
        pass
    return None

def _get_cell_style(cell):
    s = {}
    try:
        f = cell.font
        if f.bold:      s["bold"] = True
        if f.italic:    s["italic"] = True
        if f.underline: s["underline"] = True
        if f.size:      s["fontSize"] = f.size
        if f.name:      s["fontName"] = f.name
        fc = _rgb_excel(f.color)
        if fc:          s["color"] = fc
    except Exception:
        pass
    try:
        fi = cell.fill
        if fi.fill_type == "solid":
            bg = _rgb_excel(fi.fgColor)
            if bg:      s["bg"] = bg
    except Exception:
        pass
    try:
        al = cell.alignment
        if al.horizontal and al.horizontal != "general":
            s["align"] = al.horizontal
        if al.wrap_text:
            s["wrap"] = True
    except Exception:
        pass
    return s or None

_XREF_RE_EXCEL = re.compile(r"^='?([^'!]+)'?!\$?([A-Za-z]+)\$?([0-9]+)$")

def _resolve_xref_excel(formula: str, wb):
    m = _XREF_RE_EXCEL.match(formula)
    if not m:
        return None
    sheet_name, col_str, row_str = m.group(1), m.group(2), m.group(3)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
    if ws is None:
        return None
    return ws[f"{col_str.upper()}{row_str}"].value

def _parse_workbook_to_json(file_path: str) -> dict:
    import openpyxl
    from openpyxl.utils import column_index_from_string
    wb_f = openpyxl.load_workbook(file_path, data_only=False)
    wb_v = openpyxl.load_workbook(file_path, data_only=True)
    sheets_out = []
    for sheet_name in wb_f.sheetnames:
        ws_f = wb_f[sheet_name]
        ws_v = wb_v[sheet_name] if sheet_name in wb_v.sheetnames else ws_f
        cells = {}
        for row_f in ws_f.iter_rows():
            for cf in row_f:
                r0, c0 = cf.row - 1, cf.column - 1
                formula = None
                raw = None
                if isinstance(cf.value, str) and cf.value.startswith("="):
                    formula = cf.value
                    cv = ws_v.cell(cf.row, cf.column)
                    raw = cv.value
                    if raw is None:
                        raw = _resolve_xref_excel(cf.value, wb_v)
                else:
                    raw = cf.value
                if isinstance(raw, (datetime, date)) and not _is_date_format(cf.number_format or ""):
                    raw = _datetime_to_serial(raw)
                if formula is None and raw is None:
                    continue
                cd: dict = {}
                if formula:
                    cd["f"] = formula
                display = _cell_display_excel(raw, cf.number_format)
                if display:
                    cd["w"] = display
                if raw is not None:
                    if isinstance(raw, bool):
                        cd["v"] = raw; cd["t"] = "b"
                    elif isinstance(raw, (int, float)):
                        cd["v"] = raw; cd["t"] = "n"
                    elif isinstance(raw, (datetime, date)):
                        cd["v"] = raw.isoformat(); cd["t"] = "d"
                    else:
                        cd["v"] = str(raw); cd["t"] = "s"
                nf = cf.number_format
                if nf and nf not in ("General", "@", ""):
                    cd["nf"] = nf
                st = _get_cell_style(cf)
                if st:
                    cd["s"] = st
                cells[f"{r0},{c0}"] = cd
        merges = [
            [mr.min_row - 1, mr.min_col - 1, mr.max_row - 1, mr.max_col - 1]
            for mr in ws_f.merged_cells.ranges
        ]
        col_widths = {}
        for ltr, cd2 in ws_f.column_dimensions.items():
            if cd2.width:
                col_widths[str(column_index_from_string(ltr) - 1)] = round(cd2.width * 7)
        row_heights = {}
        for rn, rd in ws_f.row_dimensions.items():
            if rd.height:
                row_heights[str(rn - 1)] = round(rd.height * 4 / 3)
        sheets_out.append({
            "name": sheet_name,
            "cells": cells,
            "merges": merges,
            "colWidths": col_widths,
            "rowHeights": row_heights,
            "range": [
                (ws_f.min_row or 1) - 1,
                (ws_f.min_column or 1) - 1,
                (ws_f.max_row or 1) - 1,
                (ws_f.max_column or 1) - 1,
            ],
        })
    return {"filename": Path(file_path).name, "sheets": sheets_out}

# show the session id
@app.get("/api/agent/session")
async def agent_session():
    """Generate and return a new session ID for the agent."""
    session_id = _generate_run_session_id()
    return {"session_id": session_id}

@app.get("/api/excel/data")
async def get_excel_data(session_id: str):
    """Return openpyxl-parsed Excel data as JSON for the in-page preview."""
    info = _agent_workbook_info.get(session_id) or {}
    s3_key = info.get("s3_key")
    local_path = info.get("local")
    path = None

    # S3: download to temp file for parsing
    if s3_key and s3_enabled():
        tmp = download_to_temp(s3_key)
        if tmp and tmp.is_file():
            path = str(tmp)

    # Local fallbacks
    if not path:
        run_dir = _agent_run_dirs.get(session_id)
        if local_path and Path(local_path).is_file():
            path = local_path
        if not path and run_dir:
            xlsx_files = sorted(run_dir.glob("forecast_model_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
            if xlsx_files:
                path = str(xlsx_files[0])
        if not path:
            try:
                from config import RUNS_DIR
            except ImportError:
                RUNS_DIR = _project_root / "data" / "runs"
            fallback_dir = RUNS_DIR / session_id
            if local_path:
                fallback = fallback_dir / Path(local_path).name
                if fallback.is_file():
                    path = str(fallback)
            if not path and fallback_dir.exists():
                xlsx_files = sorted(fallback_dir.glob("forecast_model_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
                if xlsx_files:
                    path = str(xlsx_files[0])

    if not path or not Path(path).is_file():
        raise HTTPException(status_code=404, detail="Excel file not found")
    try:
        from fastapi.responses import JSONResponse
        return JSONResponse(_parse_workbook_to_json(path))
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/preview_excel")
async def render_index():
    """Excel viewer — React route /preview_excel."""
    return _spa_index()

@app.get("/api/excel")
async def download_excel(session_id: str):
    """Download the Excel workbook generated by the agent for this session."""
    info = _agent_workbook_info.get(session_id) or {}
    s3_key = info.get("s3_key")
    local_path = info.get("local")
    if s3_key and s3_enabled():
        url = get_presigned_download_url(s3_key)
        if url:
            return RedirectResponse(url=url, status_code=302)
        tmp = download_to_temp(s3_key)
        if tmp and tmp.is_file():
            return FileResponse(str(tmp), filename=tmp.name,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    path = None
    run_dir = _agent_run_dirs.get(session_id)
    if local_path and Path(local_path).is_file():
        path = local_path
    if not path and run_dir:
        xlsx_files = sorted(run_dir.glob("forecast_model_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        if xlsx_files:
            path = str(xlsx_files[0])
    if not path:
        try:
            from config import RUNS_DIR
        except ImportError:
            RUNS_DIR = _project_root / "data" / "runs"
        fallback_dir = RUNS_DIR / session_id
            
        if local_path:
            fallback = fallback_dir / Path(local_path).name
            if fallback.is_file():
                path = str(fallback)
        if not path and fallback_dir.exists():
            xlsx_files = sorted(fallback_dir.glob("forecast_model_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
            if xlsx_files:
                path = str(xlsx_files[0])
    if not path or not Path(path).is_file():
        raise HTTPException(status_code=404, detail="Excel file not found")
    return FileResponse(path, filename=Path(path).name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------------------------------------------------------------------
# PPTX agent endpoints
# ---------------------------------------------------------------------------

def _run_pptx_task(session_id: str, run_dir: Path, user_input: dict) -> None:
    """Runs in a thread pool; generates the forecast PPTX and updates _pptx_job_status."""
    _use_s3 = s3_enabled()
    try:
        from agents.pptx_agent import generate_forecast_pptx
    except ImportError as exc:
        _pptx_job_status[session_id] = {
            "status": "error",
            "error": f"pptx_agent module not found: {exc}",
        }
        return

    try:
        run_dir.mkdir(parents=True, exist_ok=True)
        template_path = str(_project_root / "agents" / "pptx_agent" / "Chrysleys PPT Template 3.pptx")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        pptx_filename = f"forecast_presentation_{ts}.pptx"
        pptx_path = str(run_dir / pptx_filename)

        generate_forecast_pptx(user_input, template_path, pptx_path)

        info: dict = {"local": pptx_path, "s3_key": None}
        if _use_s3 and upload_file:
            key = s3_output_key(session_id, pptx_filename)
            if upload_file(Path(pptx_path), key):
                info["s3_key"] = key

        _pptx_file_info[session_id] = info
        _pptx_job_status[session_id] = {
            "status": "done",
            "pptx_path": pptx_path,
            "success": True,
        }
    except Exception as exc:
        logging.getLogger("PptxEndpoint").error("PPTX error: %s", exc, exc_info=True)
        _pptx_job_status[session_id] = {
            "status": "error",
            "error": str(exc),
            "success": False,
        }


class PptxRequest(BaseModel):
    session_id: str = ""   # reuse the same session_id as the Excel agent
    user_input: dict = {}


@app.post("/api/pptx")
async def run_pptx_agent(req: PptxRequest):
    """Start PPTX generation in background; call this in parallel with /api/agent."""
    session_id = req.session_id or _generate_run_session_id()

    if _pptx_job_status.get(session_id, {}).get("status") == "running":
        return {"session_id": session_id, "status": "started", "message": "PPTX already generating"}

    if s3_enabled():
        run_dir = Path(tempfile.mkdtemp(prefix=f"fa_pptx_{session_id}_"))
    else:
        try:
            from config import RUNS_DIR
        except ImportError:
            RUNS_DIR = _project_root / "data" / "runs"
        run_dir = RUNS_DIR / session_id
    _agent_run_dirs.setdefault(session_id, run_dir)

    _pptx_job_status[session_id] = {"status": "running"}
    asyncio.create_task(asyncio.to_thread(_run_pptx_task, session_id, run_dir, req.user_input))
    return {"session_id": session_id, "status": "started", "success": True}


@app.get("/api/pptx/status")
async def pptx_status(session_id: str):
    """Poll for PPTX agent completion. Returns status: running|done|error."""
    job = _pptx_job_status.get(session_id)
    if not job:
        return {"session_id": session_id, "status": "unknown", "success": False}
    st = job.get("status", "unknown")
    out: dict = {"session_id": session_id, "status": st}
    if st == "done":
        out["success"] = True
    elif st == "error":
        out["success"] = False
        out["error"] = job.get("error", "Unknown error")
    return out


@app.get("/api/pptx")
async def download_pptx(session_id: str):
    """Download the PPTX presentation generated for this session."""
    info = _pptx_file_info.get(session_id) or {}
    s3_key = info.get("s3_key")
    local_path = info.get("local")

    if s3_key and s3_enabled():
        url = get_presigned_download_url(s3_key)
        if url:
            return RedirectResponse(url=url, status_code=302)
        tmp = download_to_temp(s3_key)
        if tmp and tmp.is_file():
            return FileResponse(
                str(tmp), filename=tmp.name,
                media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
            )

    path = None
    run_dir = _agent_run_dirs.get(session_id)
    if local_path and Path(local_path).is_file():
        path = local_path
    if not path and run_dir:
        pptx_files = sorted(
            Path(run_dir).glob("forecast_presentation_*.pptx"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        if pptx_files:
            path = str(pptx_files[0])

    if not path or not Path(path).is_file():
        raise HTTPException(status_code=404, detail="PPTX file not found")
    return FileResponse(
        path, filename=Path(path).name,
        media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
    )


# ---------------------------------------------------------------------------
# Prompt Studio endpoints
# ---------------------------------------------------------------------------
_PROMPTS_DIR = _project_root / "agents" / "excel_agent" / "prompts"
_PROMPTS_BACKUP_DIR = _PROMPTS_DIR / "backups"
_PROMPT_DELIMITER = "--------------------------"

_PROMPT_METADATA: Dict[str, dict] = {
    "flow_generate": {
        "label": "Flow Generation",
        "step": 2,
        "description": "Generates patient funnel structure from indication & product info",
        "featured": True,
        "test_type": "flow",
        "variables": ["indication", "product_name", "drug_class", "country", "launch_year", "peak_year"],
    },
    "assumptions_generate": {
        "label": "Assumption Generation",
        "step": 3,
        "description": "Extracts parameter overrides from user requirements and resources",
        "featured": True,
        "test_type": "assumptions",
        "variables": ["user_query", "default_assumptions", "indication", "product_name", "drug_class", "country", "launch_year", "peak_year", "funnel_structure"],
    },
    "action_items_generate": {
        "label": "Action Items",
        "step": None,
        "description": "Creates Excel sheet implementation plan",
        "featured": False,
        "test_type": "generic",
        "variables": ["user_query", "assumptions_config", "sheet_definition", "sheet_index"],
    },
    "decomposer_decompose_request": {
        "label": "Decomposer",
        "step": None,
        "description": "Breaks down user requests into structured components",
        "featured": False,
        "test_type": "generic",
        "variables": ["user_request"],
    },
    "critic_review_spreadsheet": {
        "label": "Critic Review",
        "step": None,
        "description": "Reviews generated spreadsheets for quality",
        "featured": False,
        "test_type": "generic",
        "variables": ["sheet_under_review", "sheet_action_plan", "sheet_content"],
    },
    "coder_generate_sheet": {
        "label": "Coder – Generate",
        "step": None,
        "description": "Generates Python/openpyxl code for an Excel sheet",
        "featured": False,
        "test_type": "generic",
        "variables": ["action_plan", "current_sheet_name"],
    },
    "coder_fix_code": {
        "label": "Coder – Fix",
        "step": None,
        "description": "Fixes errors in generated code",
        "featured": False,
        "test_type": "generic",
        "variables": ["action_plan", "current_code", "error_message"],
    },
    "selector_next_node": {
        "label": "Selector (Router)",
        "step": None,
        "description": "Routes between agent nodes in the LangGraph workflow",
        "featured": False,
        "test_type": "generic",
        "variables": ["current_state"],
    },
}


def _prompt_live_path(name: str) -> Path:
    return _PROMPTS_DIR / f"{name}.txt"


def _prompt_draft_path(name: str) -> Path:
    return _PROMPTS_DIR / f"{name}_draft.txt"


def _has_draft(name: str) -> bool:
    return _prompt_draft_path(name).exists()


def _read_prompt(name: str, draft: bool = False) -> Optional[str]:
    p = _prompt_draft_path(name) if draft else _prompt_live_path(name)
    try:
        return p.read_text(encoding="utf-8")
    except FileNotFoundError:
        return None


def _split_prompt(content: str):
    if _PROMPT_DELIMITER in content:
        system, _, user = content.partition(_PROMPT_DELIMITER)
        return system.strip(), user.strip()
    return content.strip(), ""


def _write_prompt(name: str, content: str, draft: bool = False) -> None:
    p = _prompt_draft_path(name) if draft else _prompt_live_path(name)
    p.write_text(content, encoding="utf-8")


def _backup_prompt(name: str) -> str:
    _PROMPTS_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = _PROMPTS_BACKUP_DIR / f"{name}_{ts}.txt"
    live = _prompt_live_path(name)
    if live.exists():
        shutil.copy2(live, backup_path)
    return str(backup_path)


@app.get("/prompt-editor")
async def serve_prompt_studio():
    """Prompt Studio — React route /prompt-editor."""
    return _spa_index()


@app.get("/api/prompts/locked-schema")
async def get_locked_schema():
    """Return the server-enforced output schemas so the editor can display them as read-only."""
    return {
        "flow":        _FLOW_OUTPUT_SCHEMA,
        "assumptions": _ASSUMPTIONS_OUTPUT_SCHEMA,
    }


@app.get("/api/prompts")
async def list_prompts():
    prompts = []
    if not _PROMPTS_DIR.exists():
        return {"prompts": prompts}
    for txt in sorted(_PROMPTS_DIR.glob("*.txt")):
        name = txt.stem
        if name.endswith("_draft"):
            continue
        meta = _PROMPT_METADATA.get(name, {
            "label": name.replace("_", " ").title(),
            "step": None,
            "description": "",
            "featured": False,
            "test_type": "generic",
            "variables": [],
        })
        prompts.append({
            "name": name,
            "label": meta.get("label", name),
            "step": meta.get("step"),
            "description": meta.get("description", ""),
            "featured": meta.get("featured", False),
            "test_type": meta.get("test_type", "generic"),
            "variables": meta.get("variables", []),
            "has_draft": _has_draft(name),
            "last_modified": datetime.fromtimestamp(txt.stat().st_mtime).isoformat(),
        })
    # Sort: featured prompts by step asc, non-featured alphabetically after
    prompts.sort(key=lambda p: (
        0 if p["featured"] else 1,
        p["step"] if p["step"] is not None else 999,
        p["name"],
    ))
    return {"prompts": prompts}


@app.get("/api/prompts/{name}")
async def get_prompt(name: str):
    live_path = _prompt_live_path(name)
    live_content = _read_prompt(name)
    if live_content is None:
        raise HTTPException(status_code=404, detail=f"Prompt '{name}' not found")
    live_system, live_user = _split_prompt(live_content)
    last_mod = datetime.fromtimestamp(live_path.stat().st_mtime).isoformat() if live_path.exists() else None
    result: dict = {
        "name": name,
        "meta": _PROMPT_METADATA.get(name, {}),
        "live": {"system_prompt": live_system, "user_template": live_user},
        "has_draft": _has_draft(name),
        "draft": None,
        "last_modified": last_mod,
    }
    if _has_draft(name):
        draft_content = _read_prompt(name, draft=True)
        if draft_content:
            ds, du = _split_prompt(draft_content)
            result["draft"] = {"system_prompt": ds, "user_template": du}
    return result


class _PromptDraftBody(BaseModel):
    system_prompt: str
    user_template: str


@app.post("/api/prompts/{name}/draft")
async def save_prompt_draft(name: str, body: _PromptDraftBody):
    if not _prompt_live_path(name).exists():
        raise HTTPException(status_code=404, detail=f"Prompt '{name}' not found")
    combined = f"{body.system_prompt}\n\n{_PROMPT_DELIMITER}\n\n{body.user_template}"
    _write_prompt(name, combined, draft=True)
    return {"status": "ok", "message": f"Draft saved for '{name}'"}


@app.post("/api/prompts/{name}/publish")
async def publish_prompt(name: str):
    if not _has_draft(name):
        raise HTTPException(status_code=400, detail=f"No draft found for '{name}'")
    draft_content = _read_prompt(name, draft=True)
    if not draft_content:
        raise HTTPException(status_code=400, detail="Draft is empty")
    backup_path = _backup_prompt(name)
    _write_prompt(name, draft_content, draft=False)
    _prompt_draft_path(name).unlink(missing_ok=True)
    return {"status": "ok", "message": f"'{name}' published to live", "backup_path": backup_path}


@app.delete("/api/prompts/{name}/draft")
async def discard_prompt_draft(name: str):
    draft = _prompt_draft_path(name)
    if not draft.exists():
        raise HTTPException(status_code=404, detail=f"No draft for '{name}'")
    draft.unlink()
    return {"status": "ok", "message": f"Draft discarded for '{name}'"}


class _FlowTestBody(BaseModel):
    system_prompt: str
    user_template: str
    indication: str = ""
    product_name: str = ""
    drug_class: str = ""
    country: str = ""
    launch_year: int = 2025
    peak_year: int = 2035
    query: str = ""  # optional free-text instructions (e.g. "add a compliance rate parameter")


class _ResourceItem(BaseModel):
    type: str  # "url" | "text" | "file"
    content: str
    name: Optional[str] = None


class _AssumptionsTestBody(BaseModel):
    system_prompt: str
    user_template: str
    query: str = ""
    resources: List[_ResourceItem] = []
    # Shared product context (same as Step 1 / flow test)
    indication: str = ""
    product_name: str = ""
    drug_class: str = ""
    country: str = ""
    launch_year: int = 2025
    peak_year: int = 2035
    # Funnel structure output from Step 2 (optional)
    funnel_structure: Optional[dict] = None


class _GenericTestBody(BaseModel):
    system_prompt: str
    user_template: str
    query: str = ""


async def _fetch_url_text(url: str, char_limit: int = 10_000) -> str:
    """Fetch a URL and return clean readable text suitable for LLM context.

    Uses requests (already a project dep) + stdlib html.parser to strip
    navigation, scripts, and boilerplate so the model sees article body.
    Falls back to brute-force tag removal if the smart path yields too little.
    Returns a message string for PDFs or fetch failures rather than raising.
    """
    from html.parser import HTMLParser as _HP
    import requests as _req

    class _BodyExtractor(_HP):
        # Tags whose entire subtree is discarded
        _SKIP = {"script", "style", "nav", "header", "footer", "aside",
                 "noscript", "iframe", "form", "svg", "canvas", "menu",
                 "toolbar", "advertisement", "banner"}
        # Tags that signal a paragraph break in the output
        _BLOCK = {"p", "h1", "h2", "h3", "h4", "h5", "h6", "li", "td", "th",
                  "article", "section", "blockquote", "pre", "dt", "dd", "br"}

        def __init__(self):
            super().__init__(convert_charrefs=True)
            self._skip_depth = 0
            self._buf: list[str] = []
            self._out: list[str] = []

        def handle_starttag(self, tag, attrs):
            if tag in self._SKIP:
                self._skip_depth += 1
            if tag in self._BLOCK:
                self._flush()

        def handle_endtag(self, tag):
            if tag in self._SKIP:
                self._skip_depth = max(0, self._skip_depth - 1)
            if tag in self._BLOCK:
                self._flush()

        def handle_data(self, data):
            if self._skip_depth == 0:
                self._buf.append(data)

        def _flush(self):
            t = "".join(self._buf).strip()
            if t:
                self._out.append(t)
            self._buf = []

        def result(self) -> str:
            self._flush()
            return "\n\n".join(self._out)

    def _sync_fetch() -> str:
        try:
            resp = _req.get(
                url,
                headers={
                    "User-Agent": (
                        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/120.0.0.0 Safari/537.36"
                    ),
                    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                    "Accept-Language": "en-US,en;q=0.5",
                },
                timeout=15,
                allow_redirects=True,
            )
            resp.raise_for_status()
        except Exception as exc:
            return f"[Could not fetch {url}: {exc}]"

        ct = resp.headers.get("content-type", "").lower()
        if "pdf" in ct or url.lower().endswith(".pdf"):
            return (
                "[PDF detected — the agent cannot read PDFs via URL. "
                "Download the file and upload it directly in the Resources panel instead.]"
            )

        body = resp.text

        # Plain text / JSON / XML — no parsing needed
        if "<html" not in body[:3000].lower():
            return body[:char_limit]

        # Smart extraction: skip nav/footer/scripts, keep article body
        extractor = _BodyExtractor()
        try:
            extractor.feed(body)
        except Exception:
            pass
        text = extractor.result()
        text = re.sub(r"\n{3,}", "\n\n", text).strip()

        # Fallback if smart extraction returned too little (JS-heavy shell, etc.)
        if len(text) < 200:
            text = re.sub(r"<[^>]+>", " ", body)
            text = re.sub(r"\s+", " ", text).strip()
            if len(text) < 200:
                return (
                    "[Page returned very little text — it may require JavaScript to render. "
                    "Try pasting the content as a Note instead.]"
                )

        return text[:char_limit]

    return await asyncio.to_thread(_sync_fetch)


def _extract_json(text: str, want_array: bool = False):
    """Extract the first valid JSON object or array from a freeform string.

    Uses bracket counting so inline brackets in prose text (e.g. '[see above]')
    don't corrupt the match the way a greedy regex does.
    """
    opener, closer = ('[', ']') if want_array else ('{', '}')
    alt_opener = '{' if want_array else '['
    # Try the preferred type first, then the other
    for op, cl in [(opener, closer), (alt_opener, ']' if alt_opener == '[' else '}')]:
        depth = 0
        start = None
        for i, ch in enumerate(text):
            if ch == op:
                if depth == 0:
                    start = i
                depth += 1
            elif ch == cl:
                depth -= 1
                if depth == 0 and start is not None:
                    candidate = text[start:i + 1]
                    try:
                        return json.loads(candidate)
                    except Exception:
                        # Reset and keep scanning for next opener
                        start = None
                        depth = 0
    return None


def _bedrock_call(system_prompt: str, user_message: str, temperature: float = 0.4) -> str:
    client = BedrockClient(region=os.getenv("AWS_REGION", "us-east-1"))
    resp = client.invoke_model(
        messages=[{"role": "user", "content": user_message}],
        system_prompt=system_prompt,
        max_tokens=2048,
        temperature=temperature,
    )
    return resp.get("content", "")


# ---------------------------------------------------------------------------
# Assumptions output schema — server-enforced, always appended to user message.
# This constant is NEVER editable via the prompt studio; it is appended after
# the user template is rendered so the model always sees the correct contract.
# ---------------------------------------------------------------------------
_FLOW_OUTPUT_SCHEMA = """\
─────────────────────────────────────────────────────────────
REQUIRED OUTPUT FORMAT (enforced by server — do not alter):
Return ONLY a single valid JSON object. No markdown, no code fences, no explanation.
The object must contain exactly these fields:
{
  "indication": "<indication name>",
  "market": "<country>",
  "product": "<product name>",
  "drug_class": "<drug class / MoA>",
  "preset_match": "<standard|oncology|rare|custom>",
  "epi_type": "<prevalence|incidence>",
  "recommended_params": ["population", "<IDs from the standard list or new custom IDs you define>"],
  "param_rationale": { "<param_id>": "<one-sentence rationale>", ... },
  "custom_params": {
    "<new_snake_case_id>": {
      "label": "<human-readable display name>",
      "description": "<one sentence: what this parameter measures in the patient funnel>",
      "category": "<epi|treat|market|pricing>"
    }
  },
  "ai_recommendation_text": "<2-3 sentence summary for a non-technical reader>",
  "market_summary": "<2 sentence market opportunity summary>",
  "forecast_assumptions": {
    "launch_year": <integer>,
    "peak_year": <integer>,
    "suggested_forecast_period_years": <integer>
  }
}
Standard param IDs: population, prevalence, incidence, severity, diagnosisRate, treatmentRate,
  eligibilityCriteria, progressionRate, classShare, peakProductShare, annualCostPerPatient, discount.
Rules:
  - recommended_params must start with "population" and include exactly one of "prevalence"/"incidence".
  - Do NOT force a preset template — compose the right param set for this specific indication.
  - If this indication needs parameters beyond the standard list, define them in custom_params
    with a unique snake_case ID and include that ID in recommended_params.
  - Set preset_match to "custom" whenever you deviate from the standard presets.
  - custom_params may be omitted or set to {} if no new parameters are needed.
─────────────────────────────────────────────────────────────"""

_ASSUMPTIONS_OUTPUT_SCHEMA = """\
─────────────────────────────────────────────────────────────
REQUIRED OUTPUT FORMAT (enforced by server — do not alter):
Return ONLY a valid JSON array. No markdown, no code fences, no explanation.
Each element must contain exactly these four fields:
[
  {
    "variable_name": "<dot-separated key path, e.g. general_defaults.forecast_period>",
    "value": <number or string — correct type, no quotes around numbers>,
    "rationale": "<one concise sentence explaining why this value was chosen>",
    "source": "<exact URL if derived from a provided resource, otherwise the string self>"
  }
]
Return [] only if the user explicitly says to use defaults AND no resources are provided.
─────────────────────────────────────────────────────────────"""


def _safe_template_sub(template: str, **kwargs) -> str:
    """Replace {var} placeholders for known variables only.

    Unlike str.format(), this won't raise on literal JSON braces like {"key": "value"}
    in the template body — it only substitutes the exact variable names passed as kwargs.
    """
    result = template
    for key, value in kwargs.items():
        result = result.replace("{" + key + "}", str(value))
    return result


@app.post("/api/prompts/test/flow")
async def test_flow_prompt(body: _FlowTestBody):
    try:
        fmt_vars = dict(
            indication=body.indication,
            product_name=body.product_name,
            drug_class=body.drug_class,
            country=body.country,
            launch_year=body.launch_year,
            peak_year=body.peak_year,
        )
        user_msg = _safe_template_sub(body.user_template, **fmt_vars)
        if body.query.strip():
            user_msg = user_msg + "\n\nADDITIONAL INSTRUCTIONS:\n" + body.query.strip()
        user_msg = user_msg + "\n\n" + _FLOW_OUTPUT_SCHEMA

        raw = await asyncio.to_thread(_bedrock_call, body.system_prompt, user_msg)
        parsed = _extract_json(raw, want_array=False)
        return {"status": "ok", "raw": raw, "parsed": parsed}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.post("/api/prompts/test/assumptions")
async def test_assumptions_prompt(body: _AssumptionsTestBody):
    _default_assumptions_sample = {
        "general_defaults": {"forecast_period": 10, "granularity": "yearly"},
        "population_defaults": {
            "base_population": {"value": 332000000, "source": "census"},
            "population_growth_rate": 0.01,
            "epidemiology": {"metric_type": "prevalence", "value": 0.002},
            "patient_funnel": {
                "diagnosis_rate": {"value": 0.70},
                "treatment_rate": {"value": 0.60},
            },
        },
        "market_defaults": {
            "class_share": {"value": 0.30, "lower_limit": 0.10, "upper_limit": 0.50},
            "product_share_within_class": {"value": 0.15},
        },
        "revenue_defaults": {
            "compliance_rate": 0.85,
            "pricing": {"gross_price": 25000, "discount_percentage": 0.20, "net_price": 20000},
        },
    }

    try:
        resource_blocks: List[str] = []
        source_identifiers: List[str] = []   # parallel list: what to put in "source" field
        for i, res in enumerate(body.resources):
            if res.type == "url":
                content = await _fetch_url_text(res.content)
                source_id = res.content
                resource_blocks.append(f"--- Resource {i+1} [URL: {res.content}] ---\n{content}")
            elif res.type == "text":
                source_id = "user_note"
                resource_blocks.append(f"--- Resource {i+1} [Note] ---\n{res.content}")
            elif res.type == "file":
                source_id = res.name or "uploaded_file"
                resource_blocks.append(f"--- Resource {i+1} [File: {res.name or 'upload'}] ---\n{res.content[:10000]}")
            else:
                source_id = "self"
            source_identifiers.append(f'  Resource {i+1}: source identifier = "{source_id}"')

        # Build injection blocks appended to user_query
        injection_parts: List[str] = []

        # 1. Product context (shared from Step 1 / flow test)
        product_lines = []
        if body.indication:   product_lines.append(f"Indication: {body.indication}")
        if body.product_name: product_lines.append(f"Product: {body.product_name}")
        if body.drug_class:   product_lines.append(f"Drug Class / MoA: {body.drug_class}")
        if body.country:      product_lines.append(f"Country: {body.country}")
        if body.launch_year:
            product_lines.append(f"Forecast Period: {body.launch_year}–{body.peak_year}")
        if product_lines:
            injection_parts.append("PRODUCT CONTEXT:\n" + "\n".join(product_lines))

        # 2. Funnel structure from Step 2 (if provided) — handles both new and legacy formats
        if body.funnel_structure:
            fs = body.funnel_structure
            rec_params = fs.get("recommended_params", [])
            if rec_params:
                # New format: param-ID-based recommendation
                param_labels = ", ".join(rec_params)
                epi_type     = fs.get("epi_type", "prevalence")
                preset       = fs.get("preset_match", "standard")
                ai_text      = fs.get("ai_recommendation_text", "")
                funnel_block = (
                    f"FORECAST FUNNEL PARAMETERS (from Flow Agent):\n"
                    f"  Preset: {preset}  |  Epi type: {epi_type}\n"
                    f"  Parameters: {param_labels}"
                )
                if ai_text:
                    funnel_block += f"\n  Agent note: {ai_text}"
                injection_parts.append(funnel_block)
            else:
                # Legacy funnel_stages format
                stages = fs.get("funnel_stages", [])
                if stages:
                    stage_lines = "\n".join(
                        f"  {i+1}. {s.get('label', s.get('id',''))} — "
                        f"typical: {s.get('typical_value', s.get('typical_range','?'))}"
                        for i, s in enumerate(stages)
                    )
                    injection_parts.append(f"FORECAST FUNNEL STRUCTURE:\n{stage_lines}")
                else:
                    injection_parts.append(
                        "FORECAST FUNNEL STRUCTURE:\n" + json.dumps(fs, indent=2)
                    )

        # 3. External resources — inject content or explicit "none" fallback signal
        if resource_blocks:
            source_index = "RESOURCE SOURCE IDENTIFIERS (use these exact strings in the 'source' field):\n" \
                           + "\n".join(source_identifiers)
            injection_parts.append(
                source_index
                + "\n\nADDITIONAL CONTEXT FROM RESOURCES — extract values from these first, "
                  "then use domain knowledge only for anything not covered:\n\n"
                + "\n\n".join(resource_blocks)
            )
        else:
            injection_parts.append(
                "RESOURCES PROVIDED: None — use your pharmaceutical domain knowledge "
                "to suggest appropriate assumption values based on the product context above."
            )

        enriched_query = body.query
        if injection_parts:
            enriched_query = (body.query + "\n\n" + "\n\n".join(injection_parts)).strip()

        user_msg = _safe_template_sub(
            body.user_template,
            user_query=enriched_query,
            default_assumptions=json.dumps(_default_assumptions_sample, indent=2),
            indication=body.indication,
            product_name=body.product_name,
            drug_class=body.drug_class,
            country=body.country,
            launch_year=str(body.launch_year),
            peak_year=str(body.peak_year),
            funnel_structure=json.dumps(body.funnel_structure, indent=2) if body.funnel_structure else "Not provided",
        )

        # Always append the locked output schema — this cannot be edited away via the prompt studio.
        user_msg = user_msg + "\n\n" + _ASSUMPTIONS_OUTPUT_SCHEMA

        raw = await asyncio.to_thread(_bedrock_call, body.system_prompt, user_msg)

        parsed = _extract_json(raw, want_array=True)

        readable: List[dict] = []
        if isinstance(parsed, list):
            for item in parsed:
                readable.append({
                    "key": item.get("variable_name") or item.get("key_path", ""),
                    "value": item.get("value", ""),
                    "rationale": item.get("rationale", ""),
                    "source": item.get("source", "self"),
                })

        # Build CSV string
        import csv, io
        csv_buf = io.StringIO()
        writer = csv.writer(csv_buf)
        writer.writerow(["variable_name", "value", "rationale", "source"])
        for r in readable:
            writer.writerow([r["key"], r["value"], r["rationale"], r["source"]])
        csv_text = csv_buf.getvalue()

        return {"status": "ok", "raw": raw, "parsed": parsed, "readable": readable, "csv": csv_text}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.post("/api/prompts/test/generic")
async def test_generic_prompt(body: _GenericTestBody):
    try:
        try:
            user_msg = body.user_template.format(query=body.query, user_query=body.query, user_request=body.query)
        except KeyError:
            user_msg = f"{body.user_template}\n\nINPUT:\n{body.query}"
        raw = await asyncio.to_thread(_bedrock_call, body.system_prompt, user_msg)
        parsed = _extract_json(raw, want_array=False) or _extract_json(raw, want_array=True)
        return {"status": "ok", "raw": raw, "parsed": parsed}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


def _extract_upload_text(raw_bytes: bytes, filename: str, char_limit: int = 10_000) -> str:
    """Return readable text from uploaded file bytes.

    Handles PDF via pdfminer.six if available, plain text otherwise.
    Detects binary garbage and returns a helpful message rather than junk.
    """
    fname_lower = (filename or "").lower()
    is_pdf = fname_lower.endswith(".pdf") or raw_bytes[:4] == b"%PDF"

    if is_pdf:
        try:
            from pdfminer.high_level import extract_text as _pdf_extract
            import io as _io
            text = _pdf_extract(_io.BytesIO(raw_bytes))
            text = re.sub(r"\s+", " ", text).strip()
            if not text:
                return "[PDF had no extractable text — it may be a scanned image. Try pasting key figures as a Note.]"
            return text[:char_limit]
        except ImportError:
            return (
                "[PDF extraction requires pdfminer.six.\n"
                "Run: pip install pdfminer.six\n"
                "Until then, copy-paste the relevant text as a Note in the Resources panel.]"
            )
        except Exception as exc:
            return f"[Could not read PDF: {exc}]"

    # Plain text decode (UTF-8 with latin-1 fallback)
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            text = raw_bytes.decode(enc)
            break
        except Exception:
            text = None
    else:
        text = raw_bytes.decode("utf-8", errors="ignore")

    # Sanity-check: if more than 30 % of chars are non-printable it's binary
    if text:
        non_print = sum(1 for c in text[:500] if not c.isprintable() and c not in "\n\r\t")
        if non_print / max(len(text[:500]), 1) > 0.3:
            return "[Binary file — cannot read as text. Upload a .txt, .csv, or .json file instead.]"

    return (text or "")[:char_limit]


@app.post("/api/prompts/upload-resource")
async def upload_prompt_resource(file: UploadFile = File(...)):
    raw_bytes = await file.read()
    content = await asyncio.to_thread(_extract_upload_text, raw_bytes, file.filename or "")
    return {"name": file.filename, "content": content, "size": len(raw_bytes)}


# Mount Vite build assets (JS/CSS bundles)
if (_FRONTEND_DIST / "assets").is_dir():
    app.mount("/assets", StaticFiles(directory=_FRONTEND_DIST / "assets"), name="frontend-assets")


# ---------------------------------------------------------------------------
# Dev entry-point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import uvicorn
    # PurePath.match() matches from the right, so patterns must match the
    # trailing components of the absolute path.
    # e.g. "data/runs/*/scripts/*.py" matches
    #   .../data/runs/20260403_.../scripts/sheet_1_inputs.py
    uvicorn.run(
        "forecast_server:app",
        host="0.0.0.0",
        port=8000,
        reload=True,
        reload_excludes=[
            "data/runs/*/scripts/*.py",   # generated agent scripts
            "data/runs/*/*/*.py",          # any .py nested 2-deep under a run
            "data/runs/*/*.py",            # any .py directly under a run
            "data/output/*",
            "data/logs/*",
        ],
    )
